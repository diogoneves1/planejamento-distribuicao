import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
from datetime import date
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURAÇÃO
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Liberacao de Pedidos v3.2 IA",
    layout="wide",
    initial_sidebar_state="expanded"
)

VERSION = "3.2"

# ──────────────────────────────────────────────────────────────────────────────
# ESTILOS
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"], .stApp {
    font-family: 'Inter', system-ui, -apple-system, sans-serif !important;
    font-size: 13px;
    color: #1C2B4A;
    -webkit-font-smoothing: antialiased;
}
.stApp { background-color: #F4F6FA; }

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background-color: #FFFFFF !important;
    border-right: 1px solid #E8ECF2 !important;
}
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3,
section[data-testid="stSidebar"] h4 {
    color: #1C2B4A !important;
    font-size: 10.5px !important;
    font-weight: 600 !important;
    text-transform: uppercase;
    letter-spacing: 0.09em;
    margin: 0 !important;
}
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] .stMarkdown p {
    color: #5A6A8A !important;
    font-size: 11.5px !important;
    line-height: 1.55;
}
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] p,
section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] label,
section[data-testid="stSidebar"] .stWidgetLabel p {
    color: #1C2B4A !important;
    font-size: 11.5px !important;
    font-weight: 500 !important;
}
section[data-testid="stSidebar"] div[role="radiogroup"] label p,
section[data-testid="stSidebar"] div[role="radiogroup"] label span,
section[data-testid="stSidebar"] div[role="radiogroup"] label {
    color: #1C2B4A !important;
    font-size: 12px !important;
    font-weight: 400 !important;
}
section[data-testid="stSidebar"] input {
    background: #FFFFFF !important;
    color: #1C2B4A !important;
    border: 1px solid #C8D0E4 !important;
    border-radius: 6px !important;
    font-size: 12px !important;
}
section[data-testid="stSidebar"] input:focus {
    border-color: #0054A6 !important;
    box-shadow: 0 0 0 3px rgba(0,84,166,0.12) !important;
}
section[data-testid="stSidebar"] input::placeholder { color: #9CAABE !important; }
section[data-testid="stSidebar"] div[data-baseweb="select"] > div {
    background: #FFFFFF !important;
    border: 1px solid #C8D0E4 !important;
    border-radius: 6px !important;
    color: #1C2B4A !important;
}
section[data-testid="stSidebar"] div[data-baseweb="select"] > div:focus-within {
    border-color: #0054A6 !important;
    box-shadow: 0 0 0 3px rgba(0,84,166,0.12) !important;
}
section[data-testid="stSidebar"] div[data-baseweb="select"] input {
    color: #1C2B4A !important;
    background: transparent !important;
    border: none !important;
}
section[data-testid="stSidebar"] span[data-baseweb="tag"] {
    background-color: #EBF2FF !important;
    border: 1px solid #BDD3F5 !important;
    border-radius: 4px !important;
}
section[data-testid="stSidebar"] span[data-baseweb="tag"] span {
    color: #003D80 !important;
    font-size: 11px !important;
    font-weight: 500 !important;
}
section[data-testid="stSidebar"] [data-baseweb="tag"] svg { fill: #5A8AC8 !important; }
section[data-testid="stSidebar"] hr {
    border-color: #E8ECF2 !important;
    margin: 14px 0 !important;
}
section[data-testid="stSidebar"] [data-testid="stFileUploader"] {
    background: #F7F9FC !important;
    border: 1.5px dashed #C8D0E4 !important;
    border-radius: 8px !important;
}
section[data-testid="stSidebar"] [data-testid="stFileUploader"] p,
section[data-testid="stSidebar"] [data-testid="stFileUploader"] span,
section[data-testid="stSidebar"] [data-testid="stFileUploader"] small {
    color: #5A6A8A !important;
    font-size: 11.5px !important;
}
section[data-testid="stSidebar"] .stAlert {
    background: #0054A6 !important;
    border: 1px solid #003D80 !important;
    border-radius: 6px !important;
    font-size: 11.5px !important;
    font-weight: 600 !important;
}
section[data-testid="stSidebar"] .stAlert p,
section[data-testid="stSidebar"] .stAlert span,
section[data-testid="stSidebar"] .stAlert div { color: #FFFFFF !important; }
section[data-testid="stSidebar"] .stAlert svg { fill: #FFFFFF !important; }

/* ── Botões ── */
.stButton > button {
    background-color: #0054A6 !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 6px !important;
    font-weight: 600 !important;
    font-size: 12.5px !important;
    padding: 9px 22px !important;
    box-shadow: 0 1px 4px rgba(0,84,166,0.25) !important;
    transition: background-color 0.15s !important;
}
.stButton > button:hover { background-color: #003D80 !important; }
.stDownloadButton > button {
    background-color: #FFFFFF !important;
    color: #0054A6 !important;
    border: 1.5px solid #0054A6 !important;
    border-radius: 6px !important;
    font-weight: 600 !important;
    font-size: 12.5px !important;
    padding: 9px 22px !important;
}
.stDownloadButton > button:hover { background-color: #EBF2FF !important; }

/* ── Métricas ── */
[data-testid="stMetric"] {
    background: #FFFFFF !important;
    border-radius: 8px !important;
    padding: 16px 18px !important;
    border: 1px solid #E8ECF2 !important;
    border-top: 3px solid #0054A6 !important;
    box-shadow: 0 1px 4px rgba(28,43,74,0.06) !important;
}
[data-testid="stMetric"] label,
[data-testid="stMetric"] [data-testid="stMetricLabel"] p {
    color: #5A6A8A !important;
    font-size: 10px !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.08em !important;
}
[data-testid="stMetric"] [data-testid="stMetricValue"] {
    color: #1C2B4A !important;
    font-size: 1.65rem !important;
    font-weight: 700 !important;
    line-height: 1.15 !important;
}

/* ── Abas ── */
.stTabs [data-baseweb="tab-list"] {
    background: #FFFFFF !important;
    border-bottom: 2px solid #E8ECF2 !important;
    padding: 0 !important;
    gap: 0 !important;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    border-radius: 0 !important;
    font-size: 12.5px !important;
    font-weight: 500 !important;
    color: #5A6A8A !important;
    padding: 11px 18px !important;
    border-bottom: 2px solid transparent !important;
    margin-bottom: -2px !important;
}
.stTabs [aria-selected="true"] {
    color: #0054A6 !important;
    font-weight: 600 !important;
    border-bottom: 2px solid #0054A6 !important;
}
.stTabs [data-baseweb="tab"]:hover { color: #0054A6 !important; background: #F4F8FF !important; }

/* ── Tabelas / Expander / Hr ── */
[data-testid="stDataFrame"] {
    border-radius: 8px !important;
    border: 1px solid #E8ECF2 !important;
    overflow: hidden !important;
    box-shadow: 0 1px 4px rgba(28,43,74,0.05) !important;
}
[data-testid="stExpander"] {
    background: #FFFFFF !important;
    border: 1px solid #E8ECF2 !important;
    border-radius: 8px !important;
}
hr { border-color: #E8ECF2 !important; margin: 16px 0 !important; }
.stAlert { border-radius: 6px !important; font-size: 12.5px !important; }
.stCaption, [data-testid="stCaptionContainer"] p {
    color: #5A6A8A !important;
    font-size: 11px !important;
}

/* ── Cards de gráfico ── */
.g-card {
    background: #FFFFFF;
    border: 1px solid #E8ECF2;
    border-radius: 8px;
    padding: 16px 18px 10px 18px;
    margin-bottom: 12px;
    box-shadow: 0 1px 4px rgba(28,43,74,0.05);
}
.g-label {
    font-size: 10px;
    font-weight: 700;
    color: #5A6A8A;
    text-transform: uppercase;
    letter-spacing: 0.09em;
    margin-bottom: 10px;
    padding-bottom: 8px;
    border-bottom: 1px solid #F0F3F8;
}

/* ── Topbar ── */
.topbar {
    background: #0054A6;
    padding: 0 28px;
    height: 52px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin: -1rem -1rem 0 -1rem;
    border-bottom: 1px solid #003D80;
}
.topbar-left { display: flex; align-items: center; gap: 16px; }
.topbar-logo { font-size: 14px; font-weight: 700; color: #FFFFFF; }
.topbar-logo span { font-weight: 300; opacity: 0.7; }
.topbar-divider { width: 1px; height: 20px; background: rgba(255,255,255,0.25); }
.topbar-page { font-size: 13px; font-weight: 500; color: rgba(255,255,255,0.88); }
.topbar-right { font-size: 11px; color: rgba(255,255,255,0.5); }
.topbar-badge {
    background: rgba(255,255,255,0.15);
    color: #FFFFFF;
    font-size: 10px;
    font-weight: 600;
    padding: 2px 8px;
    border-radius: 10px;
    letter-spacing: 0.05em;
}

/* ── Breadcrumb ── */
.breadcrumb {
    font-size: 11px; color: #9CAABE;
    margin: 14px 0 6px 0;
    display: flex; align-items: center; gap: 6px;
}
.breadcrumb-sep { color: #C8D0E4; }
.breadcrumb-current { color: #1C2B4A; font-weight: 500; }

/* ── Sidebar brand ── */
.sb-brand {
    display: flex; align-items: center; gap: 10px;
    padding: 16px 0 14px 0;
    border-bottom: 1px solid #E8ECF2;
    margin-bottom: 16px;
}
.sb-brand-icon {
    width: 32px; height: 32px;
    background: #0054A6; border-radius: 6px;
    display: flex; align-items: center; justify-content: center;
    font-size: 13px; color: #FFFFFF; font-weight: 700; flex-shrink: 0;
}
.sb-brand-name { font-size: 12.5px; font-weight: 600; color: #1C2B4A; line-height: 1.2; }
.sb-brand-sub  { font-size: 10px; color: #5A6A8A; margin-top: 1px; }
.sb-section {
    font-size: 10px; font-weight: 600; color: #9CAABE;
    text-transform: uppercase; letter-spacing: 0.11em;
    margin: 16px 0 7px 0;
}

/* ── Painel KPI ── */
.kpi-panel {
    background: #FFFFFF; border: 1px solid #E8ECF2;
    border-radius: 8px; padding: 16px 18px 12px 18px;
    margin-bottom: 16px;
    box-shadow: 0 1px 4px rgba(28,43,74,0.05);
}
.kpi-panel-title {
    font-size: 10px; font-weight: 700; color: #5A6A8A;
    text-transform: uppercase; letter-spacing: 0.09em;
    margin-bottom: 12px; padding-bottom: 8px;
    border-bottom: 1px solid #F0F3F8;
}
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ──────────────────────────────────────────────────────────────────────────────
COLUNAS_OBRIGATORIAS = [
    "PEDIDO", "CLIENTE", "ESTADO", "REGIÃO",
    "FATURAR EM", "ITEM", "QNTD PROGRAMADA", "ESTOQUE INICIAL"
]

C_BLUE  = "#0054A6"
C_LGRAY = "#E8ECF2"
C_RED   = "#C0392B"
C_AMBER = "#B85C00"
C_GREEN = "#217A3C"
C_TEXT  = "#1C2B4A"
C_MUTED = "#5A6A8A"


def br(v: float, decimais: int = 2) -> str:
    fmt = f"{v:,.{decimais}f}"
    return fmt.replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_peso(kg: float) -> str:
    if kg < 1000:
        return br(kg, 0) + " kg"
    ton = kg / 1000
    return f"{ton:,.2f} ton".replace(",", ".")


# ──────────────────────────────────────────────────────────────────────────────
# FUNÇÕES DE NEGÓCIO
# ──────────────────────────────────────────────────────────────────────────────

def preparar_base(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.str.strip().str.upper()
    df.columns = [c.replace("RESTRIÇÃO?", "RESTRICAO").replace("Ã", "A").replace("Ç", "C")
                  if "RESTRI" in c else c for c in df.columns]
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].astype(str).str.strip()
    df["FATURAR EM"]      = pd.to_datetime(df["FATURAR EM"], errors="coerce")
    df["QNTD PROGRAMADA"] = pd.to_numeric(df["QNTD PROGRAMADA"], errors="coerce").fillna(0)
    df["ESTOQUE INICIAL"] = pd.to_numeric(df["ESTOQUE INICIAL"], errors="coerce").fillna(0)
    if "PESO" in df.columns:
        df["PESO"] = pd.to_numeric(df["PESO"], errors="coerce").fillna(0)
    if "RESTRICAO" not in df.columns:
        df["RESTRICAO"] = "NAO"
    else:
        df["RESTRICAO"] = (df["RESTRICAO"]
            .str.strip().str.upper()
            .str.replace("Ã", "A", regex=False)
            .str.replace("Ç", "C", regex=False)
            .str.replace("?", "", regex=False)
        )
        df["RESTRICAO"] = df["RESTRICAO"].replace({
            "NAO": "NAO", "SIM": "SIM", "AJUSTAR BASE": "Ajustar Base",
        }).fillna("NAO")
    if "STATUS" not in df.columns:
        df["STATUS"] = "Nao informado"
    else:
        df["STATUS"] = df["STATUS"].fillna("Nao informado")
    return df


def validar_colunas(df: pd.DataFrame) -> list:
    return [c for c in COLUNAS_OBRIGATORIAS if c not in df.columns]


def montar_estoque(df: pd.DataFrame) -> dict:
    estoque = {}
    for _, row in df.iterrows():
        if row["ITEM"] not in estoque:
            estoque[row["ITEM"]] = row["ESTOQUE INICIAL"]
    return estoque


def aplicar_filtros(df, regioes, estados, clientes, restricoes, pedidos_filtro):
    if regioes:        df = df[df["REGIÃO"].isin(regioes)]
    if estados:        df = df[df["ESTADO"].isin(estados)]
    if clientes:       df = df[df["CLIENTE"].isin(clientes)]
    if restricoes:     df = df[df["RESTRICAO"].isin(restricoes)]
    if pedidos_filtro: df = df[df["PEDIDO"].astype(str).isin([str(p) for p in pedidos_filtro])]
    return df


def ordenar_prioridade(df, p_pedido, p_restricao, p_cliente, p_regiao, p_estado):
    df = df.copy()
    pedidos_p0 = set(df[df["PEDIDO"].astype(str).isin([str(p) for p in p_pedido])]["PEDIDO"].unique()) if p_pedido else set()
    pedidos_p1 = set(df[df["RESTRICAO"].isin(p_restricao) & ~df["PEDIDO"].isin(pedidos_p0)]["PEDIDO"].unique()) if p_restricao else set()
    pedidos_p2 = set(df[df["CLIENTE"].isin(p_cliente) & ~df["PEDIDO"].isin(pedidos_p0) & ~df["PEDIDO"].isin(pedidos_p1)]["PEDIDO"].unique()) if p_cliente else set()
    pedidos_p3 = set(df[df["REGIÃO"].isin(p_regiao) & ~df["PEDIDO"].isin(pedidos_p0) & ~df["PEDIDO"].isin(pedidos_p1) & ~df["PEDIDO"].isin(pedidos_p2)]["PEDIDO"].unique()) if p_regiao else set()
    pedidos_p4 = set(df[df["ESTADO"].isin(p_estado) & ~df["PEDIDO"].isin(pedidos_p0) & ~df["PEDIDO"].isin(pedidos_p1) & ~df["PEDIDO"].isin(pedidos_p2) & ~df["PEDIDO"].isin(pedidos_p3)]["PEDIDO"].unique()) if p_estado else set()
    df["ORDEM_PRIORIDADE"] = 5
    df.loc[df["PEDIDO"].isin(pedidos_p0), "ORDEM_PRIORIDADE"] = 0
    df.loc[df["PEDIDO"].isin(pedidos_p1), "ORDEM_PRIORIDADE"] = 1
    df.loc[df["PEDIDO"].isin(pedidos_p2), "ORDEM_PRIORIDADE"] = 2
    df.loc[df["PEDIDO"].isin(pedidos_p3), "ORDEM_PRIORIDADE"] = 3
    df.loc[df["PEDIDO"].isin(pedidos_p4), "ORDEM_PRIORIDADE"] = 4
    return df.sort_values(["ORDEM_PRIORIDADE", "FATURAR EM", "PEDIDO"], ascending=True)


def analisar_pedidos(df_ordem, estoque, tem_peso: bool):
    liberados, bloqueados = [], []
    consumo_item, faltas_item = {}, {}

    for pedido, grupo in df_ordem.groupby("PEDIDO", sort=False):
        info   = grupo.iloc[0]
        faltas = []
        pode   = True

        for _, row in grupo.iterrows():
            item  = row["ITEM"]
            qtd   = row["QNTD PROGRAMADA"]
            saldo = estoque.get(item, 0)
            if saldo < qtd:
                pode = False
                faltas.append({"ITEM": item, "QTD": qtd, "SALDO": saldo, "FALTA": qtd - saldo})

        peso_pedido = round(grupo["PESO"].sum(), 4) if tem_peso else 0

        base = {
            "PEDIDO":     pedido,
            "CLIENTE":    info["CLIENTE"],
            "ESTADO":     info["ESTADO"],
            "REGIÃO":     info["REGIÃO"],
            "RESTRICAO":  info.get("RESTRICAO", "NAO"),
            "STATUS":     info.get("STATUS", "Nao informado"),
            "FATURAR EM": info["FATURAR EM"].date() if pd.notnull(info["FATURAR EM"]) else None,
            "PRIORIDADE": info["ORDEM_PRIORIDADE"],
            "PESO_TON":   peso_pedido,
        }

        if pode:
            itens = []
            for _, row in grupo.iterrows():
                item, qtd = row["ITEM"], row["QNTD PROGRAMADA"]
                estoque[item] = estoque.get(item, 0) - qtd
                consumo_item[item] = consumo_item.get(item, 0) + qtd
                itens.append(f"{item} ({int(qtd)})")
            liberados.append({**base, "ITENS": " | ".join(itens)})
        else:
            for i, f in enumerate(faltas, 1):
                base[f"ITEM_{i}"]               = f["ITEM"]
                base[f"QTD_NECESSARIA_{i}"]     = f["QTD"]
                base[f"ESTOQUE_DISPONIVEL_{i}"] = f["SALDO"]
                base[f"FALTA_{i}"]              = f["FALTA"]
                faltas_item[f["ITEM"]] = faltas_item.get(f["ITEM"], 0) + f["FALTA"]
            bloqueados.append(base)

    _ec = pd.DataFrame(columns=["ITEM", "CONSUMO_TOTAL"])
    _ef = pd.DataFrame(columns=["ITEM", "FALTA_TOTAL"])
    df_cons = pd.DataFrame([{"ITEM": k, "CONSUMO_TOTAL": v} for k, v in consumo_item.items()]
                           ).sort_values("CONSUMO_TOTAL", ascending=False) if consumo_item else _ec
    df_falt = pd.DataFrame([{"ITEM": k, "FALTA_TOTAL": v}   for k, v in faltas_item.items()]
                           ).sort_values("FALTA_TOTAL",   ascending=False) if faltas_item else _ef
    df_est  = pd.DataFrame([{"ITEM": k, "ESTOQUE_FINAL": v}  for k, v in estoque.items()]
                           ).sort_values("ESTOQUE_FINAL",  ascending=False)
    return pd.DataFrame(liberados), pd.DataFrame(bloqueados), df_cons, df_falt, df_est


def calcular_resumo(df_lib, df_bloq, df_cons, df_falt):
    total    = len(df_lib) + len(df_bloq)
    pct      = round(len(df_lib) / total * 100, 1) if total else 0
    ton_lib  = round(df_lib["PESO_TON"].sum(),  2) if not df_lib.empty  and "PESO_TON" in df_lib.columns  else 0
    ton_bloq = round(df_bloq["PESO_TON"].sum(), 2) if not df_bloq.empty and "PESO_TON" in df_bloq.columns else 0
    return pd.DataFrame([
        {"METRICA": "Total pedidos analisados", "VALOR": total},
        {"METRICA": "Pedidos liberados",         "VALOR": len(df_lib)},
        {"METRICA": "Pedidos bloqueados",         "VALOR": len(df_bloq)},
        {"METRICA": "Taxa de liberacao (%)",      "VALOR": pct},
        {"METRICA": "Peso liberado (kg)",         "VALOR": ton_lib},
        {"METRICA": "Peso retido (kg)",           "VALOR": ton_bloq},
        {"METRICA": "Total itens consumidos",     "VALOR": int(df_cons["CONSUMO_TOTAL"].sum()) if not df_cons.empty else 0},
        {"METRICA": "SKUs em falta",              "VALOR": len(df_falt)},
    ])


def resumo_por_estado(df_lib, df_bloq):
    vals_restricao = []
    for df_tmp in [df_lib, df_bloq]:
        if not df_tmp.empty and "RESTRICAO" in df_tmp.columns:
            vals_restricao += df_tmp["RESTRICAO"].dropna().unique().tolist()
    vals_restricao = sorted(set(vals_restricao))

    rows = []
    estados = sorted(set(
        (df_lib["ESTADO"].tolist()  if not df_lib.empty  else []) +
        (df_bloq["ESTADO"].tolist() if not df_bloq.empty else [])
    ))
    for estado in estados:
        lib_e  = df_lib[df_lib["ESTADO"]   == estado] if not df_lib.empty  else pd.DataFrame()
        bloq_e = df_bloq[df_bloq["ESTADO"] == estado] if not df_bloq.empty else pd.DataFrame()
        row = {
            "ESTADO":         estado,
            "PED_LIBERADOS":  len(lib_e),
            "PED_BLOQUEADOS": len(bloq_e),
            "TON_LIBERADAS":  round(lib_e["PESO_TON"].sum(),  2) if not lib_e.empty  else 0,
            "TON_RETIDAS":    round(bloq_e["PESO_TON"].sum(), 2) if not bloq_e.empty else 0,
        }
        for val in vals_restricao:
            row[f"REST_{val}"] = int((lib_e["RESTRICAO"] == val).sum()) if not lib_e.empty and "RESTRICAO" in lib_e.columns else 0
        rows.append(row)

    df = pd.DataFrame(rows).fillna(0)
    if not df.empty:
        df["TOTAL_PEDIDOS"] = df["PED_LIBERADOS"] + df["PED_BLOQUEADOS"]
        df["TOTAL_TON"]     = df["TON_LIBERADAS"] + df["TON_RETIDAS"]
        df["TX_LIBERACAO"]  = (df["PED_LIBERADOS"] / df["TOTAL_PEDIDOS"] * 100).round(1)
        df = df.sort_values("TOTAL_TON", ascending=False)
    return df


def resumo_por_status(df_lib):
    if df_lib.empty or "STATUS" not in df_lib.columns:
        return pd.DataFrame(columns=["STATUS", "PEDIDOS", "PESO_KG"])
    grp = df_lib.groupby("STATUS").agg(
        PEDIDOS=("PEDIDO", "count"),
        PESO_KG=("PESO_TON", "sum")
    ).reset_index().sort_values("PEDIDOS", ascending=False)
    grp["PESO_KG"] = grp["PESO_KG"].round(2)
    return grp


def gerar_analise_itens(df_bloq: pd.DataFrame):
    """
    Analisa pedidos bloqueados: quantos seriam liberados repondo 1, 2 ou 3 itens.
    v3.2: inclui QTD NECESSARIA por item no detalhe e QTD TOTAL por item no ranking.
    Retorna (df_ranking, df_detalhe, df_resumo) ou (None, None, None).
    """
    from itertools import combinations

    if df_bloq.empty:
        return None, None, None

    def norm_item(v):
        if pd.isna(v): return None
        try:
            f = float(v)
            return str(int(f)) if f == int(f) else str(v)
        except:
            return str(v).strip()

    # ── MUDANÇA 1: adicionar pedido_qtds para armazenar QTD_NECESSARIA por item ──
    pedido_itens = {}   # pedido → set de itens em falta
    pedido_qtds  = {}   # pedido → {item: qtd_necessaria}
    pedido_info  = {}

    for _, row in df_bloq.iterrows():
        pedido = row["PEDIDO"]
        itens_falta = set()
        qtds_falta  = {}
        for i in range(1, 15):
            item  = norm_item(row.get(f"ITEM_{i}"))
            falta = row.get(f"FALTA_{i}", 0)
            qtd   = row.get(f"QTD_NECESSARIA_{i}", falta)
            if item and pd.notna(falta) and float(falta) > 0:
                itens_falta.add(item)
                # Usa QTD_NECESSARIA se válida e > 0; senão cai para FALTA
                qtd_val = float(qtd) if (pd.notna(qtd) and float(qtd) > 0) else float(falta)
                qtds_falta[item] = qtd_val
        if itens_falta:
            pedido_itens[pedido] = itens_falta
            pedido_qtds[pedido]  = qtds_falta
            pedido_info[pedido]  = {
                "ESTADO":    row.get("ESTADO", ""),
                "REGIAO":    row.get("REGIÃO", ""),
                "RESTRICAO": row.get("RESTRICAO", ""),
                "STATUS":    row.get("STATUS", ""),
                "PESO_KG":   round(float(row.get("PESO_TON", 0) or 0), 2),
            }

    if not pedido_itens:
        return None, None, None

    total = len(pedido_itens)
    item_pedidos = {}
    for pedido, itens in pedido_itens.items():
        for item in itens:
            item_pedidos.setdefault(item, set()).add(pedido)

    todos_itens = list(item_pedidos.keys())
    rows_det = []

    # ── MUDANÇA 2: cenários com QTD ITEM N intercalado ──

    # Cenário 1 item
    for item in todos_itens:
        liberados = [p for p in item_pedidos[item] if pedido_itens[p] - {item} == set()]
        for p in liberados:
            info = pedido_info[p]
            rows_det.append({
                "CENARIO":   "1 item",
                "QTD ITENS": 1,
                "ITEM 1":    item,
                "QTD ITEM 1": int(pedido_qtds[p].get(item, 0)),
                "ITEM 2":    "",
                "QTD ITEM 2": "",
                "ITEM 3":    "",
                "QTD ITEM 3": "",
                "PEDIDO":    p,
                "ESTADO":    info["ESTADO"],
                "REGIAO":    info["REGIAO"],
                "RESTRICAO": info["RESTRICAO"],
                "STATUS":    info["STATUS"],
                "PESO_KG":   info["PESO_KG"],
            })

    # Cenário 2 itens
    # Pedido entra só se precisa EXATAMENTE desses 2 itens (não menos, não mais)
    for ia, ib in combinations(todos_itens, 2):
        liberados = [p for p in (item_pedidos[ia] | item_pedidos[ib])
                     if pedido_itens[p] == {ia, ib}]
        for p in liberados:
            info = pedido_info[p]
            rows_det.append({
                "CENARIO":   "2 itens",
                "QTD ITENS": 2,
                "ITEM 1":    ia,
                "QTD ITEM 1": int(pedido_qtds[p].get(ia, 0)),
                "ITEM 2":    ib,
                "QTD ITEM 2": int(pedido_qtds[p].get(ib, 0)),
                "ITEM 3":    "",
                "QTD ITEM 3": "",
                "PEDIDO":    p,
                "ESTADO":    info["ESTADO"],
                "REGIAO":    info["REGIAO"],
                "RESTRICAO": info["RESTRICAO"],
                "STATUS":    info["STATUS"],
                "PESO_KG":   info["PESO_KG"],
            })

    # Cenário 3 itens
    # Pedido entra só se precisa EXATAMENTE desses 3 itens (não menos, não mais)
    for ia, ib, ic in combinations(todos_itens, 3):
        liberados = [p for p in (item_pedidos[ia] | item_pedidos[ib] | item_pedidos[ic])
                     if pedido_itens[p] == {ia, ib, ic}]
        for p in liberados:
            info = pedido_info[p]
            rows_det.append({
                "CENARIO":   "3 itens",
                "QTD ITENS": 3,
                "ITEM 1":    ia,
                "QTD ITEM 1": int(pedido_qtds[p].get(ia, 0)),
                "ITEM 2":    ib,
                "QTD ITEM 2": int(pedido_qtds[p].get(ib, 0)),
                "ITEM 3":    ic,
                "QTD ITEM 3": int(pedido_qtds[p].get(ic, 0)),
                "PEDIDO":    p,
                "ESTADO":    info["ESTADO"],
                "REGIAO":    info["REGIAO"],
                "RESTRICAO": info["RESTRICAO"],
                "STATUS":    info["STATUS"],
                "PESO_KG":   info["PESO_KG"],
            })

    if not rows_det:
        return None, None, None

    df_det = (pd.DataFrame(rows_det)
                .drop_duplicates(subset=["CENARIO", "ITEM 1", "ITEM 2", "ITEM 3", "PEDIDO"])
                .sort_values(["QTD ITENS", "ITEM 1", "ITEM 2", "ITEM 3", "ESTADO"]))

    # Reordenar colunas do detalhe: ITEM N | QTD ITEM N intercalados
    cols_det = [
        "CENARIO", "QTD ITENS",
        "ITEM 1", "QTD ITEM 1",
        "ITEM 2", "QTD ITEM 2",
        "ITEM 3", "QTD ITEM 3",
        "PEDIDO", "ESTADO", "REGIAO", "RESTRICAO", "STATUS", "PESO_KG",
    ]
    df_det = df_det[[c for c in cols_det if c in df_det.columns]]

    # ── MUDANÇA 3: ranking com QTD TOTAL ITEM N ──
    rows_res = []
    for (cenario, i1, i2, i3, n), grp in df_det.groupby(
            ["CENARIO", "ITEM 1", "ITEM 2", "ITEM 3", "QTD ITENS"]):
        estados = ", ".join(sorted(grp["ESTADO"].unique()))

        def _soma_qtd(col):
            if col not in grp.columns:
                return ""
            s = int(pd.to_numeric(grp[col], errors="coerce").fillna(0).sum())
            return s  # sempre retorna o valor, mesmo que 0

        rows_res.append({
            "CENARIO":          cenario,
            "ITEM 1":           i1,
            "QTD TOTAL ITEM 1": _soma_qtd("QTD ITEM 1") if i1 else "",
            "ITEM 2":           i2,
            "QTD TOTAL ITEM 2": _soma_qtd("QTD ITEM 2") if i2 else "",
            "ITEM 3":           i3,
            "QTD TOTAL ITEM 3": _soma_qtd("QTD ITEM 3") if i3 else "",
            "QTD ITENS":        n,
            "PEDIDOS LIBERADOS": len(grp),
            "% DO TOTAL":       f"{round(len(grp)/total*100,1)}%",
            "PESO TOTAL (kg)":  round(grp["PESO_KG"].sum(), 2),
            "ESTADOS":          estados,
        })

    df_rank = (pd.DataFrame(rows_res)
                 .sort_values(["QTD ITENS", "PEDIDOS LIBERADOS"], ascending=[True, False]))

    # Reordenar colunas do ranking: ITEM N | QTD TOTAL ITEM N intercalados
    cols_rank = [
        "CENARIO",
        "ITEM 1", "QTD TOTAL ITEM 1",
        "ITEM 2", "QTD TOTAL ITEM 2",
        "ITEM 3", "QTD TOTAL ITEM 3",
        "QTD ITENS", "PEDIDOS LIBERADOS", "% DO TOTAL", "PESO TOTAL (kg)", "ESTADOS",
    ]
    df_rank = df_rank[[c for c in cols_rank if c in df_rank.columns]]

    peso_total = sum(v["PESO_KG"] for v in pedido_info.values())
    b1 = df_rank[df_rank["QTD ITENS"]==1].iloc[0] if not df_rank[df_rank["QTD ITENS"]==1].empty else None
    b2 = df_rank[df_rank["QTD ITENS"]==2].iloc[0] if not df_rank[df_rank["QTD ITENS"]==2].empty else None
    b3 = df_rank[df_rank["QTD ITENS"]==3].iloc[0] if not df_rank[df_rank["QTD ITENS"]==3].empty else None

    def combo(r):
        if r is None: return "-"
        return " + ".join(x for x in [r["ITEM 1"], r["ITEM 2"], r["ITEM 3"]] if x)

    rows_sum = [
        ("Total pedidos bloqueados",     total),
        ("Peso total retido (kg)",        round(peso_total, 2)),
        ("Itens unicos em falta",         len(todos_itens)),
        ("", ""),
        ("REPOR 1 ITEM",                  ""),
        ("Melhor item",                   combo(b1)),
        ("Pedidos liberados",             int(b1["PEDIDOS LIBERADOS"]) if b1 is not None else 0),
        ("Peso liberado (kg)",            b1["PESO TOTAL (kg)"] if b1 is not None else 0),
        ("Estados atendidos",             b1["ESTADOS"] if b1 is not None else "-"),
        ("", ""),
        ("REPOR 2 ITENS",                 ""),
        ("Melhor combo",                  combo(b2)),
        ("Pedidos liberados",             int(b2["PEDIDOS LIBERADOS"]) if b2 is not None else 0),
        ("Peso liberado (kg)",            b2["PESO TOTAL (kg)"] if b2 is not None else 0),
        ("Estados atendidos",             b2["ESTADOS"] if b2 is not None else "-"),
        ("", ""),
        ("REPOR 3 ITENS",                 ""),
        ("Melhor combo",                  combo(b3)),
        ("Pedidos liberados",             int(b3["PEDIDOS LIBERADOS"]) if b3 is not None else 0),
        ("Peso liberado (kg)",            b3["PESO TOTAL (kg)"] if b3 is not None else 0),
        ("Estados atendidos",             b3["ESTADOS"] if b3 is not None else "-"),
    ]
    df_sum = pd.DataFrame(rows_sum, columns=["METRICA", "VALOR"])

    return df_rank, df_det, df_sum




# ──────────────────────────────────────────────────────────────────────────────
# FUNÇÕES IA — Ollama local
# ──────────────────────────────────────────────────────────────────────────────


# ──────────────────────────────────────────────────────────────────────────────
# FUNÇÕES ROTEIRIZAÇÃO
# ──────────────────────────────────────────────────────────────────────────────

# Paleta de cores por status (para a fila de pedidos na roteirizacao)
CORES_STATUS = {
    "Aprovado para Faturar":                         "#217A3C",
    "Aguardando Liberação - Com Estoque Parcial":    "#B85C00",
    "Aguardando Liberação - Sem Estoque":            "#C0392B",
    "Bloqueado - Restrição Comercial":               "#8B5CF6",
    "Em Análise":                                    "#0891B2",
    "Nao informado":                                 "#5A6A8A",
}


def cor_status(status: str) -> str:
    """Retorna a cor associada a um status, ou cinza padrao."""
    return CORES_STATUS.get(str(status), "#5A6A8A")


def preparar_pedidos_rota(df_lib: pd.DataFrame, df_base: pd.DataFrame) -> pd.DataFrame:
    """
    Monta a lista de pedidos liberados para roteirizacao.
    Caixas = soma de QNTD PROGRAMADA de todos os itens do pedido.
    Peso = PESO_TON (em kg) ja calculado.
    """
    if df_lib.empty:
        return pd.DataFrame(columns=["PEDIDO", "CLIENTE", "ESTADO", "STATUS", "PESO_KG", "CAIXAS"])

    # Somar caixas por pedido a partir da base
    caixas_por_pedido = (
        df_base.groupby("PEDIDO")["QNTD PROGRAMADA"].sum().to_dict()
        if "QNTD PROGRAMADA" in df_base.columns else {}
    )

    rows = []
    for _, row in df_lib.iterrows():
        pedido = row["PEDIDO"]
        rows.append({
            "PEDIDO":  pedido,
            "CLIENTE": row.get("CLIENTE", ""),
            "ESTADO":  row.get("ESTADO", ""),
            "STATUS":  row.get("STATUS", "Nao informado"),
            "PESO_KG": round(float(row.get("PESO_TON", 0) or 0), 2),
            "CAIXAS":  int(caixas_por_pedido.get(pedido, 0)),
        })

    df = pd.DataFrame(rows).sort_values(["ESTADO", "PEDIDO"]).reset_index(drop=True)
    return df


def gerar_excel_roteirizacao(veiculos: dict, alocacoes: dict, pedidos_rota: pd.DataFrame) -> bytes:
    """
    Gera o Excel da roteirizacao.
    veiculos: {id: {"nome":..., "capacidade_kg":..., "data":...}}
    alocacoes: {id_veiculo: [{"PEDIDO":..., "CAIXAS":..., "PESO_KG":..., ...}, ...]}
    """
    from openpyxl import Workbook

    wb = Workbook()
    del wb["Sheet"]

    f_titulo = PatternFill("solid", fgColor="002D6B")
    f_navy   = PatternFill("solid", fgColor="0054A6")
    f_cinza  = PatternFill("solid", fgColor="F4F6FA")
    f_branco = PatternFill("solid", fgColor="FFFFFF")
    f_sep    = PatternFill("solid", fgColor="E8ECF2")

    fn_title = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    fn_hdr   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    fn_norm  = Font(name="Calibri", size=10, color="1C2B4A")
    fn_tot   = Font(name="Calibri", bold=True, size=11, color="0054A6")
    fn_sub   = Font(name="Calibri", italic=True, size=9, color="5A6A8A")

    centro = Alignment(horizontal="center", vertical="center")
    esq    = Alignment(horizontal="left",   vertical="center")
    dir_   = Alignment(horizontal="right",  vertical="center")

    # ── Aba RESUMO ──
    ws = wb.create_sheet("RESUMO")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for col, w in zip("BCDEFG", [22, 16, 12, 12, 18, 16]):
        ws.column_dimensions[col].width = w

    ws.row_dimensions[2].height = 28
    ws.merge_cells("B2:G2")
    c = ws.cell(2, 2); c.value = "ROTEIRIZACAO — RESUMO DAS CARGAS"
    c.font = fn_title; c.fill = f_titulo; c.alignment = esq
    for col in range(2, 8): ws.cell(2, col).fill = f_titulo

    ws.row_dimensions[4].height = 16
    for j, lbl in enumerate(["VEICULO", "DATA CARREG.", "CAIXAS", "PESO (kg)", "CAPACIDADE (kg)", "% OCUPACAO"], 2):
        c = ws.cell(4, j); c.value = lbl; c.font = fn_hdr; c.fill = f_navy; c.alignment = centro

    linha = 5
    tot_caixas = 0
    tot_peso   = 0
    for vid, vinfo in veiculos.items():
        pedidos_v = alocacoes.get(vid, [])
        caixas_v  = sum(p["CAIXAS"] for p in pedidos_v)
        peso_v    = round(sum(p["PESO_KG"] for p in pedidos_v), 2)
        cap_v     = vinfo.get("capacidade_kg", 0) or 0
        ocup      = round(peso_v / cap_v * 100, 1) if cap_v else 0
        tot_caixas += caixas_v
        tot_peso   += peso_v

        fill = f_cinza if (linha % 2 == 0) else f_branco
        vals = [
            vinfo.get("nome", f"Veiculo {vid}"),
            vinfo.get("data", "-"),
            caixas_v,
            peso_v,
            cap_v,
            f"{ocup}%",
        ]
        for j, val in enumerate(vals, 2):
            c = ws.cell(linha, j); c.value = val; c.font = fn_norm; c.fill = fill
            c.alignment = esq if j == 2 else (centro if j in (3, 7) else dir_)
        ws.row_dimensions[linha].height = 15
        linha += 1

    # Total
    for j in range(2, 8): ws.cell(linha, j).fill = f_sep
    ws.cell(linha, 2).value = "TOTAL"; ws.cell(linha, 2).font = fn_tot; ws.cell(linha, 2).alignment = esq
    ws.cell(linha, 4).value = tot_caixas; ws.cell(linha, 4).font = fn_tot; ws.cell(linha, 4).alignment = centro
    ws.cell(linha, 5).value = round(tot_peso, 2); ws.cell(linha, 5).font = fn_tot; ws.cell(linha, 5).alignment = dir_
    ws.row_dimensions[linha].height = 16
    ws.freeze_panes = "B5"

    # ── Uma aba por veiculo ──
    for vid, vinfo in veiculos.items():
        pedidos_v = alocacoes.get(vid, [])
        if not pedidos_v:
            continue
        nome_aba = str(vinfo.get("nome", f"Veiculo {vid}"))[:28].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(nome_aba)
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 2
        for col, w in zip("BCDEF", [16, 32, 10, 12, 14]):
            ws.column_dimensions[col].width = w

        ws.row_dimensions[2].height = 28
        ws.merge_cells("B2:F2")
        c = ws.cell(2, 2); c.value = f"{vinfo.get('nome','Veiculo')}  |  Carreg.: {vinfo.get('data','-')}"
        c.font = fn_title; c.fill = f_titulo; c.alignment = esq
        for col in range(2, 7): ws.cell(2, col).fill = f_titulo

        cap_v  = vinfo.get("capacidade_kg", 0) or 0
        peso_v = round(sum(p["PESO_KG"] for p in pedidos_v), 2)
        cx_v   = sum(p["CAIXAS"] for p in pedidos_v)
        ws.row_dimensions[3].height = 14
        ws.merge_cells("B3:F3")
        c = ws.cell(3, 2)
        c.value = f"{len(pedidos_v)} pedidos  ·  {cx_v} caixas  ·  {peso_v} kg  ·  capacidade {cap_v} kg"
        c.font = fn_sub; c.alignment = esq

        ws.row_dimensions[5].height = 16
        for j, lbl in enumerate(["PEDIDO", "CLIENTE", "ESTADO", "CAIXAS", "PESO (kg)"], 2):
            c = ws.cell(5, j); c.value = lbl; c.font = fn_hdr; c.fill = f_navy; c.alignment = centro

        r = 6
        for i, p in enumerate(pedidos_v):
            fill = f_cinza if i % 2 == 0 else f_branco
            vals = [p["PEDIDO"], p.get("CLIENTE", ""), p.get("ESTADO", ""), p["CAIXAS"], p["PESO_KG"]]
            for j, val in enumerate(vals, 2):
                c = ws.cell(r, j); c.value = val; c.font = fn_norm; c.fill = fill
                c.alignment = dir_ if j in (5, 6) else (centro if j in (4, 4) else esq)
            ws.row_dimensions[r].height = 15
            r += 1

        for j in range(2, 7): ws.cell(r, j).fill = f_sep
        ws.cell(r, 2).value = "TOTAL"; ws.cell(r, 2).font = fn_tot; ws.cell(r, 2).alignment = esq
        ws.cell(r, 5).value = cx_v;   ws.cell(r, 5).font = fn_tot; ws.cell(r, 5).alignment = centro
        ws.cell(r, 6).value = peso_v; ws.cell(r, 6).font = fn_tot; ws.cell(r, 6).alignment = dir_
        ws.freeze_panes = "B6"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def montar_prompt_ia(df_lib, df_bloq, df_falt, df_cons, df_estado, df_status):
    import pandas as pd

    total      = len(df_lib) + len(df_bloq)
    liberados  = len(df_lib)
    bloqueados = len(df_bloq)
    taxa       = round(liberados / total * 100, 1) if total else 0

    # ── Peso ──
    ton_lib  = round(df_lib["PESO_TON"].sum() / 1000,  2) if not df_lib.empty  and "PESO_TON" in df_lib.columns else 0
    ton_bloq = round(df_bloq["PESO_TON"].sum() / 1000, 2) if not df_bloq.empty and "PESO_TON" in df_bloq.columns else 0

    # ── Itens em falta — top 15 ──
    top_faltas = df_falt.head(15).to_string(index=False) if not df_falt.empty else "Nenhum item em falta."

    # ── Estados — todos ──
    top_estados = df_estado.to_string(index=False) if not df_estado.empty else "Sem dados por estado."

    # ── Status ──
    top_status = df_status.to_string(index=False) if not df_status.empty else "Sem dados por status."

    # ── Consumo top 10 ──
    top_cons = df_cons.head(10).to_string(index=False) if not df_cons.empty else "Sem dados de consumo."

    # ── Cruzamento: bloqueados por restricao ──
    if not df_bloq.empty and "RESTRICAO" in df_bloq.columns:
        bloq_restricao = df_bloq.groupby("RESTRICAO")["PEDIDO"].count().reset_index()
        bloq_restricao.columns = ["RESTRICAO", "PEDIDOS_BLOQUEADOS"]
        bloq_restricao_str = bloq_restricao.to_string(index=False)
    else:
        bloq_restricao_str = "Sem dados."

    # ── Cruzamento: liberados por restricao ──
    if not df_lib.empty and "RESTRICAO" in df_lib.columns:
        lib_restricao = df_lib.groupby("RESTRICAO")["PEDIDO"].count().reset_index()
        lib_restricao.columns = ["RESTRICAO", "PEDIDOS_LIBERADOS"]
        lib_restricao_str = lib_restricao.to_string(index=False)
    else:
        lib_restricao_str = "Sem dados."

    # ── Cruzamento: peso retido por estado (top 8) ──
    if not df_estado.empty and "TON_RETIDAS" in df_estado.columns:
        peso_retido = df_estado[["ESTADO","TON_RETIDAS","PED_BLOQUEADOS"]].sort_values("TON_RETIDAS", ascending=False).head(8)
        peso_retido_str = peso_retido.to_string(index=False)
    else:
        peso_retido_str = "Sem dados."

    # ── Cruzamento: estados com taxa de liberacao abaixo de 50% ──
    if not df_estado.empty and "TX_LIBERACAO" in df_estado.columns:
        estados_criticos = df_estado[df_estado["TX_LIBERACAO"] < 50][["ESTADO","TX_LIBERACAO","PED_BLOQUEADOS","TON_RETIDAS"]]
        estados_criticos_str = estados_criticos.to_string(index=False) if not estados_criticos.empty else "Nenhum estado abaixo de 50%."
    else:
        estados_criticos_str = "Sem dados."

    # ── Cruzamento: itens em falta x pedidos bloqueados (quantos pedidos cada item trava) ──
    if not df_bloq.empty:
        item_cols = [c for c in df_bloq.columns if c.startswith("ITEM_")]
        item_count = {}
        for col in item_cols:
            for val in df_bloq[col].dropna():
                if str(val).strip() not in ("", "nan"):
                    item_count[str(val)] = item_count.get(str(val), 0) + 1
        if item_count:
            df_item_count = pd.DataFrame(
                sorted(item_count.items(), key=lambda x: -x[1]),
                columns=["ITEM", "PEDIDOS_TRAVADOS"]
            ).head(15)
            item_count_str = df_item_count.to_string(index=False)
        else:
            item_count_str = "Sem dados."
    else:
        item_count_str = "Sem dados."

    # ── Cruzamento: status x estado (liberados) ──
    if not df_lib.empty and "STATUS" in df_lib.columns and "ESTADO" in df_lib.columns:
        status_estado = df_lib.groupby(["ESTADO","STATUS"])["PEDIDO"].count().reset_index()
        status_estado.columns = ["ESTADO","STATUS","QTDE"]
        status_estado_str = status_estado.sort_values("QTDE", ascending=False).head(15).to_string(index=False)
    else:
        status_estado_str = "Sem dados."

    prompt = (
        "Voce e um analista senior especialista em Supply Chain, PCP, logistica e planejamento de distribuicao. "
        "Analise profundamente os dados abaixo de uma simulacao de liberacao de pedidos e gere um relatorio "
        "executivo completo, rico em insights e cruzamentos de dados. "
        "Seja especifico, cite numeros, codigos de itens e nomes de estados. "
        "Responda DIRETAMENTE sem cabecalho, sem saudacao, sem data, sem formato de carta. "
        "Comece imediatamente com '## 1. Resumo Executivo'.\n\n"

        "=== VISAO GERAL ===\n"
        f"- Total de pedidos analisados: {total}\n"
        f"- Pedidos liberados: {liberados} ({taxa}%)\n"
        f"- Pedidos bloqueados: {bloqueados} ({round(100-taxa,1)}%)\n"
        f"- Toneladas liberadas: {ton_lib} ton\n"
        f"- Toneladas retidas (bloqueadas): {ton_bloq} ton\n\n"

        "=== ITENS EM FALTA (codigo | quantidade total em falta) ===\n"
        f"{top_faltas}\n\n"

        "=== ITENS QUE MAIS TRAVAM PEDIDOS (item | qtde pedidos bloqueados) ===\n"
        f"{item_count_str}\n\n"

        "=== CONSUMO DOS ITENS LIBERADOS (top 10) ===\n"
        f"{top_cons}\n\n"

        "=== DESEMPENHO POR ESTADO (liberados | bloqueados | ton lib | ton ret | tx%) ===\n"
        f"{top_estados}\n\n"

        "=== ESTADOS CRITICOS — TAXA DE LIBERACAO ABAIXO DE 50% ===\n"
        f"{estados_criticos_str}\n\n"

        "=== ESTADOS COM MAIOR PESO RETIDO ===\n"
        f"{peso_retido_str}\n\n"

        "=== PEDIDOS BLOQUEADOS POR TIPO DE RESTRICAO ===\n"
        f"{bloq_restricao_str}\n\n"

        "=== PEDIDOS LIBERADOS POR TIPO DE RESTRICAO ===\n"
        f"{lib_restricao_str}\n\n"

        "=== STATUS DOS PEDIDOS LIBERADOS ===\n"
        f"{top_status}\n\n"

        "=== STATUS X ESTADO (pedidos liberados) ===\n"
        f"{status_estado_str}\n\n"

        "Gere o relatorio com os 6 topicos abaixo. Cada topico deve ter no minimo 5-8 itens ou paragrafos, "
        "com cruzamentos de dados, numeros especificos e insights acionaveis.\n\n"

        "## 1. Resumo Executivo\n"
        "Panorama geral da operacao: taxa de liberacao, volume fisico liberado vs retido em toneladas, "
        "distribuicao regional, perfil de restricoes e status predominante. "
        "Compare o desempenho entre regioes e destaque o cenario mais critico.\n\n"

        "## 2. Analise de Gargalos e Ruptura de Estoque\n"
        "Identifique os itens com maior impacto: quantos pedidos cada item trava, "
        "qual o volume fisico associado a cada falta, quais combinacoes de itens em falta "
        "sao mais criticas. Cruce item em falta x estado mais afetado. "
        "Destaque itens com estoque zerado vs estoque parcial.\n\n"

        "## 3. Impacto Regional e por Estado\n"
        "Analise cada estado critico: taxa de liberacao, peso retido, pedidos bloqueados. "
        "Identifique quais estados concentram o maior risco operacional. "
        "Cruce restricao x estado: ha estados com alto percentual de pedidos com restricao? "
        "Qual regiao esta mais prejudicada no total?\n\n"

        "## 4. Analise de Restricoes e Status\n"
        "Avalie o impacto das restricoes (SIM, NAO, Ajustar Base) nos bloqueios. "
        "Pedidos com restricao estao travando mais ou menos que os sem restricao? "
        "Analise a distribuicao de status: o que os status revelam sobre o fluxo operacional? "
        "Ha acumulo de pedidos em algum status especifico que indica gargalo no processo?\n\n"

        "## 5. Recomendacoes Prioritarias e Plano de Acao\n"
        "Liste acoes priorizadas por impacto potencial:\n"
        "- Quais itens repor primeiro e em qual quantidade minima para desbloquear o maior numero de pedidos\n"
        "- Quais estados atender com urgencia\n"
        "- Como reordenar a fila de prioridade para maximizar toneladas liberadas\n"
        "- Acoes para resolver pedidos com restricao\n"
        "- Sugestoes de ajuste no planejamento de estoque\n\n"

        "## 6. Riscos e Pontos Criticos de Atencao\n"
        "Destaque os principais riscos operacionais: itens com estoque zerado sem previsao, "
        "estados com 0% de liberacao, pedidos com multiplos itens em falta simultaneamente, "
        "concentracao de risco em poucos itens criticos, impacto financeiro estimado do volume retido "
        "e risco de perda de clientes por atraso de entrega.\n\n"

        "Seja analitico, profundo e acionavel. Use os dados fornecidos. Nao invente numeros."
    )
    return prompt
def chamar_ollama(prompt: str) -> str:
    import requests
    try:
        response = requests.post(
            "http://localhost:11434/api/generate",
            json={
                "model": "gemma4:e2b-it-q4_K_M",
                "prompt": prompt,
                "stream": False,
            },
            timeout=300,
        )
        response.raise_for_status()
        return response.json().get("response", "Sem resposta do modelo.")
    except requests.exceptions.ConnectionError:
        return "**Erro:** Nao foi possivel conectar ao Ollama. Verifique se o servico esta rodando em http://localhost:11434"
    except requests.exceptions.Timeout:
        return "**Erro:** Timeout ao aguardar resposta do modelo. Tente novamente."
    except Exception as e:
        return f"**Erro inesperado:** {str(e)}"

# ──────────────────────────────────────────────────────────────────────────────
# EXCEL — helpers
# ──────────────────────────────────────────────────────────────────────────────

def _estilo_aba(ws, cor):
    for cell in ws[1]:
        cell.fill      = PatternFill("solid", fgColor=cor)
        cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(bottom=Side(style="thin", color="FFFFFF"))
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
        fc = "F4F6FA" if i % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.fill      = PatternFill("solid", fgColor=fc)
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in ws.columns:
        letra = get_column_letter(col[0].column)
        ml = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[letra].width = min(max(ml + 3, 12), 50)
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


def _preencher(ws, r1, c1, r2, c2, fill):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).fill = fill


def _mesclar(ws, r, c1, c2, valor, font, fill=None, align=None):
    to_remove = [mr for mr in list(ws.merged_cells.ranges)
                 if mr.min_row <= r <= mr.max_row and mr.min_col <= c2 and mr.max_col >= c1]
    for mr in to_remove:
        ws.unmerge_cells(str(mr))
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(r, c1)
    cell.value     = valor
    cell.font      = font
    cell.alignment = align or Alignment(horizontal="center", vertical="center")
    if fill:
        for c in range(c1, c2 + 1):
            ws.cell(r, c).fill = fill


def _visao_gerencial(wb, df_lib, df_bloq, df_cons, df_falt, df_est, df_estado, df_status, hoje):
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint

    if "VISAO_GERENCIAL" in wb.sheetnames:
        del wb["VISAO_GERENCIAL"]
    ws = wb.create_sheet("VISAO_GERENCIAL", 0)

    total     = len(df_lib) + len(df_bloq)
    pct_lib   = round(len(df_lib) / total * 100, 1) if total else 0
    skus_falt = len(df_falt)
    ton_lib   = round(df_lib["PESO_TON"].sum(),  2) if not df_lib.empty  and "PESO_TON" in df_lib.columns  else 0
    ton_bloq  = round(df_bloq["PESO_TON"].sum(), 2) if not df_bloq.empty and "PESO_TON" in df_bloq.columns else 0

    f_navy   = PatternFill("solid", fgColor="0054A6")
    f_titulo = PatternFill("solid", fgColor="002D6B")
    f_cinza  = PatternFill("solid", fgColor="F4F6FA")
    f_branco = PatternFill("solid", fgColor="FFFFFF")
    f_sep    = PatternFill("solid", fgColor="E8ECF2")

    fn_titulo = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    fn_label  = Font(name="Calibri", bold=True, color="5A6A8A", size=8)
    fn_navy   = Font(name="Calibri", bold=True, color="0054A6", size=18)
    fn_verde  = Font(name="Calibri", bold=True, color="217A3C", size=18)
    fn_azul   = Font(name="Calibri", bold=True, color="0054A6", size=18)
    fn_normal = Font(name="Calibri", size=10, color="1C2B4A")
    fn_hdr    = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    fn_azul_v = Font(name="Calibri", bold=True, color="0054A6", size=10)
    fn_verde_v= Font(name="Calibri", bold=True, color="217A3C", size=10)
    fn_data   = Font(name="Calibri", color="FFFFFF", size=9)

    centro = Alignment(horizontal="center", vertical="center")
    esq    = Alignment(horizontal="left",   vertical="center")
    dir_   = Alignment(horizontal="right",  vertical="center")

    for i, w in enumerate([1,14,2,14,2,14,2,14,2,14,2,14,1], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for ln, h in {1:5, 2:26, 3:5, 4:12, 5:24, 6:12, 7:24, 8:5,
                  9:12, 10:16, 11:16, 12:16, 13:16, 14:16, 15:5,
                  16:12, 17:16, 18:16, 19:16, 20:16, 21:16, 22:5,
                  23:5, 24:12, 25:16, 26:16, 27:16, 28:16, 29:5, 30:12}.items():
        ws.row_dimensions[ln].height = h

    _preencher(ws, 1, 1, 80, 13, f_cinza)

    _preencher(ws, 2, 1, 2, 13, f_titulo)
    _mesclar(ws, 2, 2, 9, f"  VISAO GERENCIAL — LIBERACAO DE PEDIDOS  |  v{VERSION}", fn_titulo, f_titulo,
             Alignment(horizontal="left", vertical="center"))
    c = ws.cell(2, 11); c.value = hoje.strftime("%d/%m/%Y")
    c.font = fn_data; c.alignment = Alignment(horizontal="right", vertical="center"); c.fill = f_titulo
    for col in [12, 13]: ws.cell(2, col).fill = f_titulo

    _preencher(ws, 3, 1, 3, 13, f_navy)

    # KPI linha 1
    _preencher(ws, 4, 1, 4, 13, f_branco)
    for col, lbl in [(2,"PEDIDOS ANALISADOS"),(4,"LIBERADOS"),(6,"BLOQUEADOS"),(8,"TAXA DE LIBERACAO"),(10,"SKUS EM FALTA")]:
        _mesclar(ws, 4, col, col+1, lbl, fn_label, f_branco, centro)
    _preencher(ws, 5, 1, 5, 13, f_branco)
    for col, val, fn in [(2,total,fn_navy),(4,len(df_lib),fn_verde),(6,len(df_bloq),fn_azul),(8,f"{pct_lib}%",fn_navy),(10,skus_falt,fn_azul)]:
        _mesclar(ws, 5, col, col+1, val, fn, f_branco, centro)

    # KPI linha 2
    _preencher(ws, 6, 1, 6, 13, f_branco)
    for col, lbl in [(2,"PESO LIBERADO (KG)"),(6,"PESO RETIDO (KG)"),(10,"TOTAL (KG)")]:
        _mesclar(ws, 6, col, col+3, lbl, fn_label, f_branco, centro)
    _preencher(ws, 7, 1, 7, 13, f_branco)
    tot_ton = round(ton_lib + ton_bloq, 2)
    def _br(v): return f"{v:,.2f}".replace(",","X").replace(".",",").replace("X",".")
    for col, val, fn in [(2,_br(ton_lib),fn_verde),(6,_br(ton_bloq),fn_azul),(10,_br(tot_ton),fn_navy)]:
        _mesclar(ws, 7, col, col+3, val, fn, f_branco, centro)

    _preencher(ws, 8, 1, 8, 13, f_sep)

    # Tabela por estado
    rest_cols_xls = [c for c in df_estado.columns if c.startswith("REST_")]
    _preencher(ws, 9, 1, 9, 13, f_navy); ws.row_dimensions[9].height = 13
    _mesclar(ws, 9, 2, 12, "PEDIDOS POR ESTADO", fn_hdr, f_navy, centro)
    _preencher(ws, 10, 1, 10, 13, f_navy); ws.row_dimensions[10].height = 13
    for col, lbl in [(2,"ESTADO"),(4,"LIB"),(5,"BLOQ"),(6,"TON LIB"),(8,"TON RETIDA")]:
        c = ws.cell(10, col); c.value = lbl; c.font = fn_hdr; c.alignment = centro; c.fill = f_navy
    for idx, rcol in enumerate(rest_cols_xls):
        col = 10 + idx; lbl = rcol.replace("REST_", "")
        c = ws.cell(10, col); c.value = lbl; c.font = fn_hdr; c.alignment = centro; c.fill = f_navy
    col_tx = 10 + len(rest_cols_xls)
    c = ws.cell(10, col_tx); c.value = "TX%"; c.font = fn_hdr; c.alignment = centro; c.fill = f_navy

    top_est = df_estado.head(6).reset_index(drop=True) if not df_estado.empty else pd.DataFrame()
    for i in range(6):
        r = 11 + i; ws.row_dimensions[r].height = 15
        fl = f_branco if i % 2 == 0 else PatternFill("solid", fgColor="F4F6FA")
        _preencher(ws, r, 2, r, 12, fl)
        if i < len(top_est):
            row_d = top_est.iloc[i]
            for col, val, fn, aln in [(2,row_d["ESTADO"],fn_normal,esq),(4,int(row_d["PED_LIBERADOS"]),fn_verde_v,centro),(5,int(row_d["PED_BLOQUEADOS"]),fn_azul_v,centro),(6,_br(row_d["TON_LIBERADAS"]),fn_verde_v,dir_),(8,_br(row_d["TON_RETIDAS"]),fn_azul_v,dir_)]:
                c = ws.cell(r, col); c.value = val; c.font = fn; c.alignment = aln; c.fill = fl
            for idx, rcol in enumerate(rest_cols_xls):
                col = 10 + idx; c = ws.cell(r, col)
                c.value = int(row_d.get(rcol, 0)); c.font = fn_azul_v; c.alignment = centro; c.fill = fl
            col_tx = 10 + len(rest_cols_xls); c = ws.cell(r, col_tx)
            c.value = f"{row_d['TX_LIBERACAO']}%"; c.font = fn_navy; c.alignment = centro; c.fill = fl

    _preencher(ws, 17, 1, 17, 13, f_sep)

    # Tabela por status
    _preencher(ws, 18, 1, 18, 13, f_navy); ws.row_dimensions[18].height = 13
    _mesclar(ws, 18, 2, 12, "PEDIDOS LIBERADOS POR STATUS", fn_hdr, f_navy, centro)
    _preencher(ws, 19, 1, 19, 13, f_navy); ws.row_dimensions[19].height = 13
    for col, lbl in [(2,"STATUS"),(8,"PEDIDOS"),(10,"PESO KG")]:
        c = ws.cell(19, col); c.value = lbl; c.font = fn_hdr; c.alignment = centro; c.fill = f_navy

    top_status = df_status.head(8).reset_index(drop=True) if not df_status.empty else pd.DataFrame()
    for i in range(8):
        r = 20 + i; ws.row_dimensions[r].height = 15
        fl = f_branco if i % 2 == 0 else PatternFill("solid", fgColor="F4F6FA")
        _preencher(ws, r, 2, r, 12, fl)
        if i < len(top_status):
            row_d = top_status.iloc[i]
            _mesclar(ws, r, 2, 7, str(row_d["STATUS"]), fn_normal, fl, esq)
            c = ws.cell(r, 8); c.value = int(row_d["PEDIDOS"]); c.font = fn_azul_v; c.alignment = centro; c.fill = fl
            ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
            c2 = ws.cell(r, 10); c2.value = _br(row_d["PESO_KG"]); c2.font = fn_verde_v; c2.alignment = dir_; c2.fill = fl
            ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)

    _preencher(ws, 28, 1, 28, 13, f_sep)

    # Gráficos
    ws.row_dimensions[29].height = 12
    _preencher(ws, 29, 2, 29, 12, f_navy)
    _mesclar(ws, 29, 2, 12, "GRAFICOS GERENCIAIS", fn_hdr, f_navy, centro)

    ws.cell(65, 1).value = "Liberados"; ws.cell(66, 1).value = "Bloqueados"
    ws.cell(65, 2).value = len(df_lib);  ws.cell(66, 2).value = len(df_bloq)

    pizza = PieChart(); pizza.title = "Liberados vs Bloqueados"
    pizza.style = 10; pizza.width = 10; pizza.height = 7
    pizza.add_data(Reference(ws, min_col=2, min_row=65, max_row=66))
    pizza.set_categories(Reference(ws, min_col=1, min_row=65, max_row=66))
    dp0 = DataPoint(idx=0); dp0.graphicalProperties.solidFill = "0054A6"
    dp1 = DataPoint(idx=1); dp1.graphicalProperties.solidFill = "C5D3E8"
    pizza.series[0].dPt = [dp0, dp1]
    ws.add_chart(pizza, "B30")

    if not df_estado.empty:
        est_top = df_estado.head(8)
        ws.cell(65,4).value="Estado"; ws.cell(65,5).value="Ton Lib"; ws.cell(65,6).value="Ton Ret"
        for i, rd in enumerate(est_top.itertuples()):
            ws.cell(66+i,4).value=rd.ESTADO; ws.cell(66+i,5).value=float(rd.TON_LIBERADAS); ws.cell(66+i,6).value=float(rd.TON_RETIDAS)
        n = len(est_top)
        barras = BarChart(); barras.type="col"; barras.title="Peso por Estado"
        barras.style=10; barras.width=14; barras.height=7; barras.grouping="clustered"
        barras.add_data(Reference(ws,min_col=5,min_row=65,max_row=65+n),titles_from_data=True)
        barras.add_data(Reference(ws,min_col=6,min_row=65,max_row=65+n),titles_from_data=True)
        barras.set_categories(Reference(ws,min_col=4,min_row=66,max_row=65+n))
        barras.series[0].graphicalProperties.solidFill="0054A6"
        barras.series[1].graphicalProperties.solidFill="C5D3E8"
        ws.add_chart(barras, "G30")

    for r in range(65, 85): ws.row_dimensions[r].hidden = True
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


def _escrever_aba_analise(wb, nome, df_data, titulo, cor_hdr="0054A6"):
    """Escreve uma aba da análise de itens no workbook."""
    if df_data is None or df_data.empty:
        return
    ws = wb.create_sheet(nome)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 5
    ws.row_dimensions[4].height = 16

    f_titulo = PatternFill("solid", fgColor="002D6B")
    f_hdr    = PatternFill("solid", fgColor=cor_hdr)
    f_cinza  = PatternFill("solid", fgColor="F4F6FA")
    f_branco = PatternFill("solid", fgColor="FFFFFF")
    fn_title = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    fn_hdr   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    fn_norm  = Font(name="Calibri", size=10, color="1C2B4A")
    centro   = Alignment(horizontal="center", vertical="center")
    esq      = Alignment(horizontal="left",   vertical="center")

    ncols = len(df_data.columns)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=ncols+1)
    c = ws.cell(2, 2); c.value = titulo; c.font = fn_title; c.fill = f_titulo; c.alignment = esq
    for col in range(2, ncols+2): ws.cell(2, col).fill = f_titulo

    for j, col in enumerate(df_data.columns, 2):
        c = ws.cell(4, j); c.value = col; c.font = fn_hdr; c.fill = f_hdr; c.alignment = centro

    for i, row in enumerate(df_data.itertuples(index=False)):
        r = 5 + i; ws.row_dimensions[r].height = 15
        fill = f_cinza if i % 2 == 0 else f_branco
        for j, val in enumerate(row, 2):
            c = ws.cell(r, j); c.value = val; c.font = fn_norm; c.fill = fill
            c.alignment = esq if isinstance(val, str) and len(str(val)) > 12 else centro

    for j, col in enumerate(df_data.columns, 2):
        try: ml = max(len(str(col)), df_data[col].astype(str).str.len().max())
        except: ml = len(str(col))
        ws.column_dimensions[get_column_letter(j)].width = min(max(ml+3,12), 55)

    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"B4:{get_column_letter(ncols+1)}4"


def gerar_excel(df_lib, df_bloq, df_cons, df_falt, df_est, df_resumo, df_estado, df_status,
                df_rank=None, df_det=None) -> bytes:
    from openpyxl import load_workbook
    output = BytesIO()
    SHEETS = {
        "RESUMO":        (df_resumo,  "0054A6"),
        "LIBERADOS":     (df_lib,     "217A3C"),
        "NAO_LIBERADOS": (df_bloq,    "0054A6"),
        "POR_ESTADO":    (df_estado,  "0054A6"),
        "POR_STATUS":    (df_status,  "0054A6"),
        "CONSUMO_ITEM":  (df_cons,    "0054A6"),
        "FALTAS_ITEM":   (df_falt,    "0054A6"),
        "ESTOQUE_FINAL": (df_est,     "217A3C"),
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for nome, (df, cor) in SHEETS.items():
            df.to_excel(writer, sheet_name=nome, index=False)
            _estilo_aba(writer.sheets[nome], cor)
    output.seek(0)
    wb = load_workbook(output)
    _visao_gerencial(wb, df_lib, df_bloq, df_cons, df_falt, df_est, df_estado, df_status, date.today())
    if "VISAO_GERENCIAL" in wb.sheetnames:
        idx = wb.sheetnames.index("VISAO_GERENCIAL")
        wb.move_sheet("VISAO_GERENCIAL", offset=-idx)
    if df_rank is not None:
        _escrever_aba_analise(wb, "ANALISE_RANKING",  df_rank, "RANKING DE COMBINACOES DE ITENS", "0054A6")
        _escrever_aba_analise(wb, "ANALISE_DETALHE",  df_det,  "DETALHE POR PEDIDO — FILTRAVEL",   "217A3C")
    final = BytesIO(); wb.save(final)
    return final.getvalue()


def gerar_excel_analise(df_rank, df_det, df_sum) -> bytes:
    """Gera o Excel da análise de liberação por item."""
    from openpyxl import Workbook
    wb = Workbook()
    del wb["Sheet"]

    f_cinza  = PatternFill("solid", fgColor="F4F6FA")
    f_branco = PatternFill("solid", fgColor="FFFFFF")
    f_titulo = PatternFill("solid", fgColor="002D6B")
    f_secao  = PatternFill("solid", fgColor="E8ECF2")

    fn_hdr   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    fn_norm  = Font(name="Calibri", size=10, color="1C2B4A")
    fn_dest  = Font(name="Calibri", bold=True, size=11, color="0054A6")
    fn_title = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    centro   = Alignment(horizontal="center", vertical="center")
    esq      = Alignment(horizontal="left",   vertical="center")

    def escrever_aba(nome, df_data, titulo, cor_hdr="0054A6"):
        if df_data is None or df_data.empty: return
        ws = wb.create_sheet(nome)
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 2
        ws.row_dimensions[1].height = 8
        ws.row_dimensions[2].height = 28
        ws.row_dimensions[3].height = 5
        ws.row_dimensions[4].height = 16
        ncols = len(df_data.columns)
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=ncols+1)
        c = ws.cell(2, 2); c.value = titulo; c.font = fn_title; c.fill = f_titulo; c.alignment = esq
        for col in range(2, ncols+2): ws.cell(2, col).fill = f_titulo
        f_h = PatternFill("solid", fgColor=cor_hdr)
        for j, col in enumerate(df_data.columns, 2):
            c = ws.cell(4, j); c.value = col; c.font = fn_hdr; c.fill = f_h; c.alignment = centro
        for i, row in enumerate(df_data.itertuples(index=False)):
            r = 5 + i; ws.row_dimensions[r].height = 15
            fill = f_cinza if i % 2 == 0 else f_branco
            for j, val in enumerate(row, 2):
                c = ws.cell(r, j); c.value = val; c.font = fn_norm; c.fill = fill
                c.alignment = esq if isinstance(val, str) and len(str(val)) > 12 else centro
        for j, col in enumerate(df_data.columns, 2):
            try: ml = max(len(str(col)), df_data[col].astype(str).str.len().max())
            except: ml = len(str(col))
            ws.column_dimensions[get_column_letter(j)].width = min(max(ml+3,12), 55)
        ws.freeze_panes = "B5"
        ws.auto_filter.ref = f"B4:{get_column_letter(ncols+1)}4"

    # RESUMO
    ws_r = wb.create_sheet("RESUMO", 0)
    ws_r.sheet_view.showGridLines = False
    ws_r.column_dimensions["A"].width = 2
    ws_r.column_dimensions["B"].width = 42
    ws_r.column_dimensions["C"].width = 22
    ws_r.row_dimensions[2].height = 28
    ws_r.merge_cells("B2:C2")
    c = ws_r.cell(2, 2); c.value = "ANALISE DE LIBERACAO — RESUMO EXECUTIVO"
    c.font = fn_title; c.fill = f_titulo; c.alignment = esq
    ws_r.row_dimensions[3].height = 5

    for i, (lbl, val) in enumerate(df_sum.values):
        r = 4 + i; ws_r.row_dimensions[r].height = 16
        eh_secao = str(val) == "" and str(lbl) != "" and str(lbl).isupper()
        fill = f_secao if eh_secao else (f_cinza if i % 2 == 0 else f_branco)
        c1 = ws_r.cell(r, 2); c1.value = lbl
        c1.font = Font(name="Calibri", bold=eh_secao, size=10, color="0054A6" if eh_secao else "1C2B4A")
        c1.fill = fill; c1.alignment = esq
        if str(val) != "":
            c2 = ws_r.cell(r, 3); c2.value = val; c2.font = fn_dest; c2.fill = fill; c2.alignment = centro

    escrever_aba("RANKING_COMBOS",  df_rank, "RANKING DE COMBINACOES", "0054A6")
    escrever_aba("DETALHE_PEDIDOS", df_det,  "DETALHE POR PEDIDO — FILTRAVEL", "217A3C")

    output = BytesIO(); wb.save(output)
    return output.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# GRÁFICOS
# ──────────────────────────────────────────────────────────────────────────────

_L = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="Inter, system-ui, sans-serif", size=11, color=C_TEXT),
)


def fig_donut(liberados, bloqueados):
    total = liberados + bloqueados
    pct   = round(liberados / total * 100, 1) if total else 0
    fig   = go.Figure(go.Pie(
        labels=["Liberados","Bloqueados"], values=[liberados,bloqueados], hole=0.72,
        marker=dict(colors=[C_BLUE,"#C5D3E8"], line=dict(color="#FFFFFF",width=2)),
        textinfo="none", hovertemplate="%{label}: %{value} (%{percent})<extra></extra>",
    ))
    fig.add_annotation(
        text=f"<b style='font-size:22px;color:{C_BLUE}'>{pct}%</b><br>"
             f"<span style='font-size:11px;color:{C_MUTED}'>liberado</span>",
        x=0.5, y=0.5, showarrow=False, align="center"
    )
    fig.update_layout(**_L, margin=dict(t=6,b=6,l=6,r=6), height=210, showlegend=True,
                      legend=dict(orientation="h", x=0.5, xanchor="center", y=-0.06,
                                  font=dict(size=11,color=C_MUTED)))
    return fig


def fig_ton_estado(df_estado):
    df_p = df_estado.head(10).sort_values("TON_LIBERADAS", ascending=True)
    fig  = go.Figure()
    fig.add_trace(go.Bar(name="Liberadas", y=df_p["ESTADO"], x=df_p["TON_LIBERADAS"],
                         orientation="h", marker_color=C_BLUE,
                         text=df_p["TON_LIBERADAS"].apply(fmt_peso),
                         textposition="outside", textfont=dict(size=10,color=C_TEXT)))
    fig.add_trace(go.Bar(name="Retidas", y=df_p["ESTADO"], x=df_p["TON_RETIDAS"],
                         orientation="h", marker_color="#C5D3E8",
                         text=df_p["TON_RETIDAS"].apply(fmt_peso),
                         textposition="outside", textfont=dict(size=10,color=C_MUTED)))
    fig.update_layout(**_L, barmode="group", height=max(200,len(df_p)*38),
                      margin=dict(t=6,b=6,l=6,r=80),
                      xaxis=dict(showgrid=True,gridcolor="#F0F3F8",zeroline=False,
                                 tickfont=dict(size=10,color=C_MUTED),title="Peso (kg)"),
                      yaxis=dict(showgrid=False,tickfont=dict(size=11,color=C_TEXT)),
                      legend=dict(orientation="h",x=1,xanchor="right",y=1.1,
                                  font=dict(size=11,color=C_MUTED),bgcolor="rgba(0,0,0,0)"))
    return fig


def fig_ped_estado(df_estado):
    df_p = df_estado.head(10).sort_values("PED_LIBERADOS", ascending=True)
    fig  = go.Figure()
    fig.add_trace(go.Bar(name="Liberados", y=df_p["ESTADO"], x=df_p["PED_LIBERADOS"],
                         orientation="h", marker_color=C_BLUE,
                         text=df_p["PED_LIBERADOS"], textposition="outside",
                         textfont=dict(size=10,color=C_TEXT)))
    fig.add_trace(go.Bar(name="Bloqueados", y=df_p["ESTADO"], x=df_p["PED_BLOQUEADOS"],
                         orientation="h", marker_color="#C5D3E8",
                         text=df_p["PED_BLOQUEADOS"], textposition="outside",
                         textfont=dict(size=10,color=C_MUTED)))
    fig.update_layout(**_L, barmode="group", height=max(200,len(df_p)*38),
                      margin=dict(t=6,b=6,l=6,r=60),
                      xaxis=dict(showgrid=True,gridcolor="#F0F3F8",zeroline=False,
                                 tickfont=dict(size=10,color=C_MUTED)),
                      yaxis=dict(showgrid=False,tickfont=dict(size=11,color=C_TEXT)),
                      legend=dict(orientation="h",x=1,xanchor="right",y=1.1,
                                  font=dict(size=11,color=C_MUTED),bgcolor="rgba(0,0,0,0)"))
    return fig


def fig_restricao_estado(df_estado):
    rest_cols = [c for c in df_estado.columns if c.startswith("REST_")]
    if not rest_cols: return go.Figure()
    df_p = df_estado.head(10).copy().sort_values(rest_cols[0], ascending=True)
    paleta = [C_BLUE, C_AMBER, C_GREEN, "#8B5CF6", "#0891B2", "#DB2777", "#D97706"]
    fig = go.Figure()
    for i, col in enumerate(rest_cols):
        fig.add_trace(go.Bar(
            name=col.replace("REST_",""), y=df_p["ESTADO"], x=df_p[col],
            orientation="h", marker_color=paleta[i % len(paleta)],
            text=df_p[col].astype(int), textposition="outside",
            textfont=dict(size=10,color=C_TEXT),
        ))
    fig.update_layout(**_L, barmode="group", height=max(200,len(df_p)*38),
                      margin=dict(t=6,b=6,l=6,r=60),
                      xaxis=dict(showgrid=True,gridcolor="#F0F3F8",zeroline=False,
                                 tickfont=dict(size=10,color=C_MUTED)),
                      yaxis=dict(showgrid=False,tickfont=dict(size=11,color=C_TEXT)),
                      legend=dict(orientation="h",x=1,xanchor="right",y=1.1,
                                  font=dict(size=11,color=C_MUTED),bgcolor="rgba(0,0,0,0)"))
    return fig


def fig_status(df_status):
    df_p = df_status.sort_values("PEDIDOS", ascending=True)
    fig  = go.Figure(go.Bar(
        y=df_p["STATUS"].astype(str), x=df_p["PEDIDOS"],
        orientation="h", marker_color=C_BLUE, marker_line=dict(width=0),
        text=df_p["PEDIDOS"], textposition="outside",
        textfont=dict(size=11,color=C_TEXT),
        hovertemplate="%{y}: %{x}<extra></extra>",
    ))
    fig.update_layout(**_L, height=max(200,len(df_p)*36),
                      margin=dict(t=6,b=6,l=6,r=60),
                      xaxis=dict(showgrid=True,gridcolor="#F0F3F8",zeroline=False,
                                 tickfont=dict(size=10,color=C_MUTED)),
                      yaxis=dict(showgrid=False,tickfont=dict(size=11,color=C_TEXT)))
    return fig


def fig_analise_itens(df_rank):
    if df_rank is None or df_rank.empty: return go.Figure()
    grp = df_rank.groupby("CENARIO")["PEDIDOS LIBERADOS"].max().reset_index()
    fig = go.Figure(go.Bar(
        x=grp["CENARIO"], y=grp["PEDIDOS LIBERADOS"],
        marker_color=C_BLUE, marker_line=dict(width=0),
        text=grp["PEDIDOS LIBERADOS"], textposition="outside",
        textfont=dict(size=12,color=C_TEXT),
    ))
    fig.update_layout(**_L, height=220,
                      margin=dict(t=10,b=10,l=10,r=10),
                      xaxis=dict(showgrid=False,tickfont=dict(size=11,color=C_TEXT)),
                      yaxis=dict(showgrid=True,gridcolor="#F0F3F8",zeroline=False,
                                 tickfont=dict(size=10,color=C_MUTED)))
    return fig


# ──────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ──────────────────────────────────────────────────────────────────────────────

with st.sidebar:

    st.markdown(f"""
    <div class="sb-brand">
        <div class="sb-brand-icon">LP</div>
        <div>
            <div class="sb-brand-name">Liberacao de Pedidos</div>
            <div class="sb-brand-sub">Versao {VERSION}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    arquivo = st.file_uploader("Planilha base (.xlsx)", type=["xlsx"])
    if arquivo:
        st.success("Planilha carregada.")

    st.markdown('<div class="sb-section">Periodo de Analise</div>', unsafe_allow_html=True)
    tipo_data = st.radio("Tipo", ["Data especifica", "Periodo"], label_visibility="collapsed")
    if tipo_data == "Data especifica":
        data_ref    = st.date_input("Data", value=date.today(), format="DD/MM/YYYY", label_visibility="collapsed")
        data_inicio = data_fim = data_ref
    else:
        data_inicio = st.date_input("Inicio", value=date.today(), format="DD/MM/YYYY")
        data_fim    = st.date_input("Fim",    value=date.today(), format="DD/MM/YYYY")

    opcoes_regiao    = []
    opcoes_estado    = []
    opcoes_cliente   = []
    opcoes_restricao = []
    opcoes_status    = []
    opcoes_pedidos   = []
    df_global = None
    tem_peso  = False

    if arquivo:
        try:
            df_tmp = preparar_base(pd.read_excel(arquivo))
            if not validar_colunas(df_tmp):
                df_global        = df_tmp
                opcoes_regiao    = sorted(df_tmp["REGIÃO"].dropna().unique())
                opcoes_estado    = sorted(df_tmp["ESTADO"].dropna().unique())
                opcoes_cliente   = sorted(df_tmp["CLIENTE"].dropna().unique())
                opcoes_restricao = sorted(df_tmp["RESTRICAO"].dropna().unique())
                opcoes_status    = sorted(df_tmp["STATUS"].dropna().unique())
                opcoes_pedidos   = sorted(df_tmp["PEDIDO"].dropna().astype(str).unique())
                tem_peso         = "PESO" in df_tmp.columns
        except Exception:
            pass

    st.markdown('<div class="sb-section">Prioridade de Liberacao</div>', unsafe_allow_html=True)
    st.caption("P0 = pedidos  ·  P1 = restricao  ·  P2 = clientes  ·  P3 = regioes  ·  P4 = estados  ·  P5 = demais")

    p_pedido    = st.multiselect("Pedidos prioritarios (P0)",   opcoes_pedidos,   placeholder="Nenhum")
    p_restricao = st.multiselect("Restricao prioritaria (P1)",  opcoes_restricao, placeholder="Nenhuma")
    p_cliente   = st.multiselect("Clientes prioritarios (P2)",  opcoes_cliente,   placeholder="Nenhum")
    p_regiao    = st.multiselect("Regioes prioritarias (P3)",   opcoes_regiao,    placeholder="Nenhuma")
    p_estado    = st.multiselect("Estados prioritarios (P4)",   opcoes_estado,    placeholder="Nenhum")

    st.markdown('<div class="sb-section">Filtros Operacionais</div>', unsafe_allow_html=True)
    f_pedido    = st.multiselect("Pedido",    opcoes_pedidos,   placeholder="Todos")
    f_regiao    = st.multiselect("Regiao",    opcoes_regiao,    placeholder="Todas")
    f_estado    = st.multiselect("Estado",    opcoes_estado,    placeholder="Todos")
    f_cliente   = st.multiselect("Cliente",   opcoes_cliente,   placeholder="Todos")
    f_restricao = st.multiselect("Restricao", opcoes_restricao, placeholder="Todas")

    st.markdown("---")

    if "rodar_analise" not in st.session_state:
        st.session_state.rodar_analise = False

    if st.button("Gerar Analise", use_container_width=True):
        st.session_state.rodar_analise = True

    rodar = st.session_state.rodar_analise


# ──────────────────────────────────────────────────────────────────────────────
# TOPBAR
# ──────────────────────────────────────────────────────────────────────────────

st.markdown(f"""
<div class="topbar">
    <div class="topbar-left">
        <div class="topbar-logo">Gestao <span>Operacional</span></div>
        <div class="topbar-divider"></div>
        <div class="topbar-page">Planejamento de Liberacao de Pedidos</div>
        <div class="topbar-badge">v{VERSION}</div>
    </div>
    <div class="topbar-right">{date.today().strftime('%d/%m/%Y')}</div>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="breadcrumb">
    Estoque <span class="breadcrumb-sep">›</span>
    Simulacao <span class="breadcrumb-sep">›</span>
    <span class="breadcrumb-current">Liberacao de Pedidos</span>
</div>
""", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# VALIDAÇÕES
# ──────────────────────────────────────────────────────────────────────────────

if not arquivo:
    st.info("Carregue a planilha base na barra lateral para iniciar a analise.")
    st.stop()

if df_global is None:
    erros = validar_colunas(preparar_base(pd.read_excel(arquivo)))
    st.error(f"Colunas obrigatorias ausentes: {', '.join(erros)}")
    st.caption("Esperado: " + "  ·  ".join(COLUNAS_OBRIGATORIAS))
    st.stop()

df = df_global

if not tem_peso:
    st.warning("Coluna PESO nao encontrada. Peso sera exibido como zero.")

with st.expander("Visualizar base carregada", expanded=False):
    st.dataframe(df.head(100), use_container_width=True, hide_index=True)
    peso_info = "  ·  coluna PESO detectada" if tem_peso else "  ·  sem coluna PESO"
    st.caption(
        f"{len(df):,} linhas  ·  {df['PEDIDO'].nunique():,} pedidos  "
        f"·  {df['ITEM'].nunique():,} itens  ·  {df['CLIENTE'].nunique():,} clientes{peso_info}"
    )

if not rodar:
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Linhas na base",  f"{len(df):,}")
    c2.metric("Pedidos unicos",  f"{df['PEDIDO'].nunique():,}")
    c3.metric("Itens unicos",    f"{df['ITEM'].nunique():,}")
    c4.metric("Clientes unicos", f"{df['CLIENTE'].nunique():,}")
    st.stop()


# ──────────────────────────────────────────────────────────────────────────────
# PROCESSAMENTO
# ──────────────────────────────────────────────────────────────────────────────

with st.spinner("Processando analise..."):
    df_f = df[
        (df["FATURAR EM"].dt.date >= data_inicio) &
        (df["FATURAR EM"].dt.date <= data_fim)
    ]
    if df_f.empty:
        st.warning("Nenhum pedido no periodo selecionado.")
        st.stop()

    df_f = aplicar_filtros(df_f, f_regiao, f_estado, f_cliente, f_restricao, f_pedido)
    if df_f.empty:
        st.warning("Nenhum pedido com os filtros aplicados.")
        st.stop()

    estoque  = montar_estoque(df)
    df_ordem = ordenar_prioridade(df_f, p_pedido, p_restricao, p_cliente, p_regiao, p_estado)
    df_lib, df_bloq, df_cons, df_falt, df_est = analisar_pedidos(df_ordem, estoque, tem_peso)
    df_resumo  = calcular_resumo(df_lib, df_bloq, df_cons, df_falt)
    df_estado  = resumo_por_estado(df_lib, df_bloq)
    df_status  = resumo_por_status(df_lib)

with st.spinner("Calculando analise de itens..."):
    df_rank, df_det, df_sum_analise = gerar_analise_itens(df_bloq)


# ──────────────────────────────────────────────────────────────────────────────
# KPIs
# ──────────────────────────────────────────────────────────────────────────────

total    = len(df_lib) + len(df_bloq)
pct      = round(len(df_lib) / total * 100, 1) if total else 0
ton_lib  = round(df_lib["PESO_TON"].sum(),  2) if not df_lib.empty  else 0
ton_bloq = round(df_bloq["PESO_TON"].sum(), 2) if not df_bloq.empty else 0

st.markdown('<div class="kpi-panel"><div class="kpi-panel-title">Resultado Geral da Simulacao</div>', unsafe_allow_html=True)

c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("Analisados",     f"{total:,}")
c2.metric("Liberados",      f"{len(df_lib):,}")
c3.metric("Bloqueados",     f"{len(df_bloq):,}")
c4.metric("Taxa Liberacao", f"{pct:.1f}%")
c5.metric("SKUs em Falta",  f"{len(df_falt):,}")

if tem_peso:
    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
    t1, t2, t3 = st.columns(3)
    t1.metric("Peso Liberado",   fmt_peso(ton_lib))
    t2.metric("Peso Retido",     fmt_peso(ton_bloq))
    t3.metric("Total Analisado", fmt_peso(ton_lib + ton_bloq))

st.markdown('</div>', unsafe_allow_html=True)
st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# ABAS
# ──────────────────────────────────────────────────────────────────────────────

tab_dash, tab_status, tab_estado, tab_lib, tab_bloq, tab_falt, tab_cons, tab_est, tab_analise, tab_ia, tab_rota = st.tabs([
    "Visao Gerencial",
    "Por Status",
    "Por Estado",
    "Liberados",
    "Nao Liberados",
    "Faltas por Item",
    "Consumo por Item",
    "Estoque Final",
    "Analise de Itens",
    "Analise IA",
    "Roteirizacao",
])


# ── VISÃO GERENCIAL ──────────────────────────────────────────────────────────
with tab_dash:
    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    col_a, col_b = st.columns([1, 2])
    with col_a:
        st.markdown('<div class="g-card"><div class="g-label">Taxa de Liberacao</div>', unsafe_allow_html=True)
        st.plotly_chart(fig_donut(len(df_lib), len(df_bloq)), use_container_width=True, config={"displayModeBar": False})
        st.markdown('</div>', unsafe_allow_html=True)
    with col_b:
        st.markdown('<div class="g-card"><div class="g-label">Pedidos por Estado — Liberados vs Bloqueados</div>', unsafe_allow_html=True)
        if not df_estado.empty:
            st.plotly_chart(fig_ped_estado(df_estado), use_container_width=True, config={"displayModeBar": False})
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    col_c, col_d = st.columns(2)
    with col_c:
        st.markdown('<div class="g-card"><div class="g-label">Restricao por Estado (Liberados)</div>', unsafe_allow_html=True)
        if not df_estado.empty:
            st.plotly_chart(fig_restricao_estado(df_estado), use_container_width=True, config={"displayModeBar": False})
        st.markdown('</div>', unsafe_allow_html=True)
    with col_d:
        st.markdown('<div class="g-card"><div class="g-label">Pedidos Liberados por Status</div>', unsafe_allow_html=True)
        if not df_status.empty:
            st.plotly_chart(fig_status(df_status), use_container_width=True, config={"displayModeBar": False})
        else:
            st.caption("Sem dados de status.")
        st.markdown('</div>', unsafe_allow_html=True)

    if tem_peso and not df_estado.empty:
        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
        st.markdown('<div class="g-card"><div class="g-label">Peso por Estado — Liberado vs Retido</div>', unsafe_allow_html=True)
        st.plotly_chart(fig_ton_estado(df_estado), use_container_width=True, config={"displayModeBar": False})
        st.markdown('</div>', unsafe_allow_html=True)


# ── POR STATUS ───────────────────────────────────────────────────────────────
with tab_status:
    if df_status.empty:
        st.info("Sem dados de status.")
    else:
        st.caption(f"{len(df_status):,} status distintos nos pedidos liberados")
        st.dataframe(df_status, use_container_width=True, hide_index=True)


# ── POR ESTADO ───────────────────────────────────────────────────────────────
with tab_estado:
    if df_estado.empty:
        st.info("Sem dados por estado.")
    else:
        st.caption(f"{len(df_estado):,} estados na analise")
        st.dataframe(df_estado, use_container_width=True, hide_index=True)


# ── LIBERADOS ────────────────────────────────────────────────────────────────
with tab_lib:
    if df_lib.empty:
        st.warning("Nenhum pedido foi liberado.")
    else:
        st.caption(f"{len(df_lib):,} pedidos liberados  ·  {fmt_peso(ton_lib)}")
        st.dataframe(df_lib, use_container_width=True, hide_index=True)


# ── NAO LIBERADOS ────────────────────────────────────────────────────────────
with tab_bloq:
    if df_bloq.empty:
        st.success("Todos os pedidos foram liberados.")
    else:
        st.caption(f"{len(df_bloq):,} pedidos bloqueados  ·  {fmt_peso(ton_bloq)} retidos")
        st.dataframe(df_bloq, use_container_width=True, hide_index=True)


# ── FALTAS ───────────────────────────────────────────────────────────────────
with tab_falt:
    if df_falt.empty:
        st.success("Nenhum item em falta.")
    else:
        st.caption(f"{len(df_falt):,} SKUs com falta registrada")
        st.dataframe(df_falt, use_container_width=True, hide_index=True)


# ── CONSUMO ──────────────────────────────────────────────────────────────────
with tab_cons:
    if df_cons.empty:
        st.info("Sem consumo registrado.")
    else:
        st.caption(f"{len(df_cons):,} itens consumidos")
        st.dataframe(df_cons, use_container_width=True, hide_index=True)


# ── ESTOQUE FINAL ─────────────────────────────────────────────────────────────
with tab_est:
    st.caption("Saldo simulado por item apos liberacao dos pedidos possiveis")
    st.dataframe(df_est, use_container_width=True, hide_index=True)


# ── ANALISE DE ITENS ─────────────────────────────────────────────────────────
with tab_analise:
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    if df_bloq.empty:
        st.success("Todos os pedidos foram liberados. Nenhuma analise necessaria.")
    elif df_rank is None:
        st.info("Nenhum pedido pode ser liberado repondo ate 3 itens com o estoque atual.")
    else:
        total_bloq = len(df_bloq)
        max_lib_1  = int(df_rank[df_rank["QTD ITENS"]==1]["PEDIDOS LIBERADOS"].max()) if not df_rank[df_rank["QTD ITENS"]==1].empty else 0
        max_lib_2  = int(df_rank[df_rank["QTD ITENS"]==2]["PEDIDOS LIBERADOS"].max()) if not df_rank[df_rank["QTD ITENS"]==2].empty else 0
        max_lib_3  = int(df_rank[df_rank["QTD ITENS"]==3]["PEDIDOS LIBERADOS"].max()) if not df_rank[df_rank["QTD ITENS"]==3].empty else 0

        st.markdown('<div class="kpi-panel"><div class="kpi-panel-title">Potencial de Liberacao por Reposicao de Itens</div>', unsafe_allow_html=True)
        ka, kb, kc, kd = st.columns(4)
        ka.metric("Pedidos Bloqueados", f"{total_bloq:,}")
        kb.metric("Repondo 1 item",     f"{max_lib_1:,} pedidos")
        kc.metric("Repondo 2 itens",    f"{max_lib_2:,} pedidos")
        kd.metric("Repondo 3 itens",    f"{max_lib_3:,} pedidos")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        st.markdown('<div class="g-card"><div class="g-label">Maximo de Pedidos Liberaveis por Cenario</div>', unsafe_allow_html=True)
        st.plotly_chart(fig_analise_itens(df_rank), use_container_width=True, config={"displayModeBar": False})
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        sub_rank, sub_det = st.tabs(["Ranking de Combos", "Detalhe por Pedido"])

        with sub_rank:
            st.caption(f"{len(df_rank):,} combinacoes  ·  filtre por CENARIO ou ITEM  ·  QTD TOTAL = unidades necessarias para atender todos os pedidos do combo")
            st.dataframe(df_rank, use_container_width=True, hide_index=True)

        with sub_det:
            st.caption(f"{len(df_det):,} linhas  ·  filtre por estado, restricao ou status  ·  QTD ITEM = unidades necessarias naquele pedido especifico")
            st.dataframe(df_det, use_container_width=True, hide_index=True)


# ── ANALISE IA ───────────────────────────────────────────────────────────────
with tab_ia:
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    st.markdown('<div class="kpi-panel"><div class="kpi-panel-title">Analise Executiva — Gerada com IA Local (Ollama)</div>', unsafe_allow_html=True)
    st.caption("A IA interpreta o resultado ja calculado pelo sistema e gera uma analise operacional. Nao altera nenhuma logica de liberacao.")
    st.markdown('</div>', unsafe_allow_html=True)

    if "resposta_ia" not in st.session_state:
        st.session_state.resposta_ia = ""

    if st.button("Gerar Analise IA", use_container_width=False):
        with st.spinner("Aguardando resposta do modelo local... isso pode levar alguns segundos."):
            prompt   = montar_prompt_ia(df_lib, df_bloq, df_falt, df_cons, df_estado, df_status)
            resposta = chamar_ollama(prompt)
            st.session_state.resposta_ia = resposta

    if st.session_state.resposta_ia:
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        st.markdown('<div class="g-card"><div class="g-label">Analise Gerada</div>', unsafe_allow_html=True)
        st.markdown(st.session_state.resposta_ia)
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.info("Clique em 'Gerar Analise IA' para iniciar a analise executiva com o modelo local.")


# ── ROTEIRIZACAO ─────────────────────────────────────────────────────────────
with tab_rota:
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    if df_lib.empty:
        st.info("Nenhum pedido liberado para roteirizar. Gere a analise primeiro.")
    else:
        # Preparar pedidos disponíveis
        pedidos_rota = preparar_pedidos_rota(df_lib, df)

        # ── Estado da sessão ──
        if "rota_veiculos" not in st.session_state:
            st.session_state.rota_veiculos = {}      # {id: {nome, capacidade_kg, data}}
        if "rota_alocacoes" not in st.session_state:
            st.session_state.rota_alocacoes = {}     # {id_veiculo: [pedido_dict, ...]}
        if "rota_next_id" not in st.session_state:
            st.session_state.rota_next_id = 1

        # Conjunto de pedidos já alocados (para remover da fila)
        pedidos_alocados = set()
        for lst in st.session_state.rota_alocacoes.values():
            for p in lst:
                pedidos_alocados.add(p["PEDIDO"])

        # ── Cabeçalho com KPIs ──
        total_pedidos   = len(pedidos_rota)
        pedidos_na_fila = total_pedidos - len(pedidos_alocados)
        total_veiculos  = len(st.session_state.rota_veiculos)

        st.markdown('<div class="kpi-panel"><div class="kpi-panel-title">Roteirizacao — Formacao de Cargas</div>', unsafe_allow_html=True)
        kr1, kr2, kr3, kr4 = st.columns(4)
        kr1.metric("Pedidos Liberados", f"{total_pedidos:,}")
        kr2.metric("Na Fila",           f"{pedidos_na_fila:,}")
        kr3.metric("Alocados",          f"{len(pedidos_alocados):,}")
        kr4.metric("Veiculos",          f"{total_veiculos:,}")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        # ── Layout: fila (esquerda) + veículos (direita) ──
        col_fila, col_veiculos = st.columns([1, 2])

        # ══════════════════════════════════════════════════════════
        # PAINEL ESQUERDO — FILA DE PEDIDOS
        # ══════════════════════════════════════════════════════════
        with col_fila:
            st.markdown('<div class="g-card"><div class="g-label">Fila de Pedidos Liberados</div>', unsafe_allow_html=True)

            # Filtros da fila
            estados_disp = sorted(pedidos_rota["ESTADO"].unique())
            filtro_estado_rota = st.multiselect("Filtrar estado", estados_disp, placeholder="Todos", key="rota_filtro_estado")
            busca_pedido = st.text_input("Buscar pedido", placeholder="Numero do pedido", key="rota_busca")

            # Aplicar filtros
            fila = pedidos_rota[~pedidos_rota["PEDIDO"].isin(pedidos_alocados)].copy()
            if filtro_estado_rota:
                fila = fila[fila["ESTADO"].isin(filtro_estado_rota)]
            if busca_pedido:
                fila = fila[fila["PEDIDO"].astype(str).str.contains(busca_pedido, case=False, na=False)]

            st.caption(f"{len(fila)} pedidos na fila")

            # Lista de veículos para o seletor de alocação
            opcoes_veiculos = {
                vid: vinfo.get("nome", f"Veiculo {vid}")
                for vid, vinfo in st.session_state.rota_veiculos.items()
            }

            if fila.empty:
                st.info("Nenhum pedido na fila com os filtros atuais.")
            else:
                # Agrupar por estado
                for estado in sorted(fila["ESTADO"].unique()):
                    grupo = fila[fila["ESTADO"] == estado]
                    st.markdown(f"**{estado}** · {len(grupo)} pedidos")

                    for _, ped in grupo.iterrows():
                        cor = cor_status(ped["STATUS"])
                        st.markdown(
                            f'<div style="border-left:4px solid {cor}; background:#FFFFFF; '
                            f'border:1px solid #E8ECF2; border-left:4px solid {cor}; '
                            f'border-radius:6px; padding:8px 10px; margin-bottom:6px;">'
                            f'<div style="font-weight:600; font-size:12px; color:#1C2B4A;">{ped["PEDIDO"]}</div>'
                            f'<div style="font-size:10.5px; color:#5A6A8A;">{ped["CLIENTE"][:32]}</div>'
                            f'<div style="font-size:10.5px; color:#1C2B4A; margin-top:2px;">'
                            f'{ped["CAIXAS"]} cx · {fmt_peso(ped["PESO_KG"])} · '
                            f'<span style="color:{cor};">{ped["STATUS"]}</span></div>'
                            f'</div>',
                            unsafe_allow_html=True
                        )
                        # Botão de alocação
                        if opcoes_veiculos:
                            cbtn1, cbtn2 = st.columns([2, 1])
                            with cbtn1:
                                destino = st.selectbox(
                                    "Veiculo", list(opcoes_veiculos.keys()),
                                    format_func=lambda v: opcoes_veiculos[v],
                                    key=f"dest_{ped['PEDIDO']}", label_visibility="collapsed"
                                )
                            with cbtn2:
                                if st.button("Alocar", key=f"aloc_{ped['PEDIDO']}", use_container_width=True):
                                    p_dict = {
                                        "PEDIDO":  ped["PEDIDO"],
                                        "CLIENTE": ped["CLIENTE"],
                                        "ESTADO":  ped["ESTADO"],
                                        "STATUS":  ped["STATUS"],
                                        "CAIXAS":  int(ped["CAIXAS"]),
                                        "PESO_KG": float(ped["PESO_KG"]),
                                    }
                                    st.session_state.rota_alocacoes.setdefault(destino, []).append(p_dict)
                                    st.rerun()
                        else:
                            st.caption("Crie um veiculo ao lado para alocar")

            st.markdown('</div>', unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════
        # PAINEL DIREITO — VEÍCULOS
        # ══════════════════════════════════════════════════════════
        with col_veiculos:
            # Botão adicionar veículo
            cadd1, cadd2 = st.columns([1, 3])
            with cadd1:
                if st.button("+ Adicionar Veiculo", use_container_width=True):
                    vid = st.session_state.rota_next_id
                    st.session_state.rota_veiculos[vid] = {
                        "nome": f"Veiculo {vid:02d}",
                        "capacidade_kg": 0,
                        "data": "",
                    }
                    st.session_state.rota_alocacoes[vid] = []
                    st.session_state.rota_next_id += 1
                    st.rerun()
            with cadd2:
                st.caption("Cada veiculo e um card de carga. Defina nome, capacidade e data de carregamento.")

            if not st.session_state.rota_veiculos:
                st.info("Nenhum veiculo criado. Clique em '+ Adicionar Veiculo' para comecar a montar as cargas.")
            else:
                # Renderizar veículos em grade de 2 colunas
                vids = list(st.session_state.rota_veiculos.keys())
                for i in range(0, len(vids), 2):
                    linha_vids = vids[i:i+2]
                    cols = st.columns(2)
                    for col_idx, vid in enumerate(linha_vids):
                        vinfo    = st.session_state.rota_veiculos[vid]
                        pedidos_v = st.session_state.rota_alocacoes.get(vid, [])
                        peso_v   = round(sum(p["PESO_KG"] for p in pedidos_v), 2)
                        caixas_v = sum(p["CAIXAS"] for p in pedidos_v)
                        cap_v    = vinfo.get("capacidade_kg", 0) or 0
                        ocupacao = round(peso_v / cap_v * 100, 1) if cap_v else 0

                        # Cor da barra de ocupação
                        if ocupacao == 0:      cor_barra = "#C5D3E8"
                        elif ocupacao <= 70:   cor_barra = "#217A3C"
                        elif ocupacao <= 100:  cor_barra = "#B85C00"
                        else:                  cor_barra = "#C0392B"

                        with cols[col_idx]:
                            st.markdown(
                                f'<div style="background:#FFFFFF; border:1px solid #E8ECF2; '
                                f'border-top:3px solid #0054A6; border-radius:8px; padding:12px 14px; '
                                f'margin-bottom:10px; box-shadow:0 1px 4px rgba(28,43,74,0.05);">',
                                unsafe_allow_html=True
                            )

                            # Nome editável
                            novo_nome = st.text_input(
                                "Nome", value=vinfo["nome"],
                                key=f"nome_{vid}", label_visibility="collapsed"
                            )
                            if novo_nome != vinfo["nome"]:
                                st.session_state.rota_veiculos[vid]["nome"] = novo_nome

                            # Capacidade e data
                            cc1, cc2 = st.columns(2)
                            with cc1:
                                nova_cap = st.number_input(
                                    "Capacidade (kg)", min_value=0, value=int(cap_v),
                                    step=100, key=f"cap_{vid}"
                                )
                                if nova_cap != cap_v:
                                    st.session_state.rota_veiculos[vid]["capacidade_kg"] = nova_cap
                            with cc2:
                                nova_data = st.text_input(
                                    "Data carreg.", value=vinfo.get("data", ""),
                                    placeholder="DD/MM/AAAA", key=f"data_{vid}"
                                )
                                if nova_data != vinfo.get("data", ""):
                                    st.session_state.rota_veiculos[vid]["data"] = nova_data

                            # Métricas da carga
                            st.markdown(
                                f'<div style="display:flex; justify-content:space-between; '
                                f'margin:8px 0 4px 0; font-size:11px;">'
                                f'<span style="color:#5A6A8A;">{len(pedidos_v)} pedidos</span>'
                                f'<span style="color:#1C2B4A; font-weight:600;">{caixas_v} cx</span>'
                                f'<span style="color:#1C2B4A; font-weight:600;">{fmt_peso(peso_v)}</span>'
                                f'</div>',
                                unsafe_allow_html=True
                            )

                            # Barra de ocupação
                            largura = min(ocupacao, 100)
                            aviso = ""
                            if cap_v and ocupacao > 100:
                                aviso = f'<span style="color:#C0392B; font-weight:600;"> · EXCEDIDO</span>'
                            st.markdown(
                                f'<div style="background:#F0F3F8; border-radius:4px; height:8px; '
                                f'overflow:hidden; margin-bottom:4px;">'
                                f'<div style="background:{cor_barra}; width:{largura}%; height:100%;"></div>'
                                f'</div>'
                                f'<div style="font-size:10px; color:#5A6A8A; margin-bottom:8px;">'
                                f'{ocupacao}% da capacidade{aviso}</div>',
                                unsafe_allow_html=True
                            )

                            # Lista de pedidos alocados
                            if pedidos_v:
                                for j, p in enumerate(pedidos_v):
                                    cp1, cp2 = st.columns([4, 1])
                                    with cp1:
                                        st.markdown(
                                            f'<div style="font-size:10.5px; color:#1C2B4A; padding:2px 0;">'
                                            f'{p["PEDIDO"]} · {p["ESTADO"]} · {p["CAIXAS"]}cx · {fmt_peso(p["PESO_KG"])}</div>',
                                            unsafe_allow_html=True
                                        )
                                    with cp2:
                                        if st.button("x", key=f"rem_{vid}_{p['PEDIDO']}_{j}", use_container_width=True):
                                            st.session_state.rota_alocacoes[vid].pop(j)
                                            st.rerun()
                            else:
                                st.caption("Nenhum pedido alocado")

                            # Botão remover veículo
                            if st.button("Remover Veiculo", key=f"delveic_{vid}", use_container_width=True):
                                # Devolver pedidos à fila
                                del st.session_state.rota_veiculos[vid]
                                if vid in st.session_state.rota_alocacoes:
                                    del st.session_state.rota_alocacoes[vid]
                                st.rerun()

                            st.markdown('</div>', unsafe_allow_html=True)

                # ── Exportar roteirização ──
                st.markdown("---")
                cexp1, cexp2 = st.columns([1, 2])
                with cexp1:
                    rota_bytes = gerar_excel_roteirizacao(
                        st.session_state.rota_veiculos,
                        st.session_state.rota_alocacoes,
                        pedidos_rota
                    )
                    st.download_button(
                        label="Baixar Roteirizacao",
                        data=rota_bytes,
                        file_name=f"roteirizacao_{date.today().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                with cexp2:
                    st.caption("Gera Excel com uma aba por veiculo + resumo geral das cargas")


# ──────────────────────────────────────────────────────────────────────────────
# EXPORTAR
# ──────────────────────────────────────────────────────────────────────────────

st.markdown("---")
col_dl, col_txt = st.columns([1, 3])
with col_dl:
    excel_bytes = gerar_excel(df_lib, df_bloq, df_cons, df_falt, df_est, df_resumo, df_estado, df_status, df_rank, df_det)
    st.download_button(
        label="Baixar Relatorio Excel",
        data=excel_bytes,
        file_name=f"liberacao_v{VERSION}_{date.today().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with col_txt:
    st.caption(
        f"10 abas: Visao Gerencial · Resumo · Liberados · Nao Liberados · Por Estado · Por Status · Consumo · Faltas · Estoque Final · Analise Ranking · Analise Detalhe"
        f"  ·  Gerado em {date.today().strftime('%d/%m/%Y')}"
    )
