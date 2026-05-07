import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
from datetime import date
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURAÇÃO
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Liberacao de Pedidos",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ──────────────────────────────────────────────────────────────────────────────
# TEMA — Visual corporativo moderno
#
#  Topbar        #0054A6  (azul primário)
#  Sidebar       #FFFFFF  com borda direita #E8ECF2
#  Fundo página  #F4F6FA
#  Cards         #FFFFFF  borda #E8ECF2  radius 8px
#  Primário      #0054A6
#  Primário dark #003D80
#  Fonte         Inter
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

/* ── Reset base ── */
html, body, [class*="css"], .stApp {
    font-family: 'Inter', system-ui, -apple-system, sans-serif !important;
    font-size: 13px;
    color: #1C2B4A;
    -webkit-font-smoothing: antialiased;
}
.stApp {
    background-color: #F4F6FA;
}

/* ════════════════════════════════════════════
   TOPBAR — barra azul fixa no topo
════════════════════════════════════════════ */
.totvs-topbar {
    position: sticky;
    top: 0;
    z-index: 100;
    background: #0054A6;
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 0 28px;
    height: 52px;
    border-bottom: 1px solid #003D80;
    margin: -1rem -1rem 0 -1rem;
}
.totvs-topbar-left {
    display: flex;
    align-items: center;
    gap: 16px;
}
.totvs-logo {
    font-size: 15px;
    font-weight: 700;
    color: #FFFFFF;
    letter-spacing: -0.02em;
}
.totvs-logo span {
    font-weight: 300;
    opacity: 0.75;
}
.totvs-divider {
    width: 1px;
    height: 20px;
    background: rgba(255,255,255,0.25);
}
.totvs-page-name {
    font-size: 13px;
    font-weight: 500;
    color: rgba(255,255,255,0.88);
    letter-spacing: 0.01em;
}
.totvs-topbar-right {
    font-size: 11.5px;
    color: rgba(255,255,255,0.55);
}

/* ════════════════════════════════════════════
   SIDEBAR — branca, borda direita sutil
════════════════════════════════════════════ */
section[data-testid="stSidebar"] {
    background-color: #FFFFFF !important;
    border-right: 1px solid #E8ECF2 !important;
}

/* Títulos de grupo */
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3,
section[data-testid="stSidebar"] h4 {
    color: #1C2B4A !important;
    font-size: 10.5px !important;
    font-weight: 600 !important;
    text-transform: uppercase;
    letter-spacing: 0.09em;
    margin: 0 !important;
}

/* Parágrafos */
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] .stMarkdown p {
    color: #00000 !important;
    font-size: 11.5px !important;
    line-height: 1.55;
}

/* ── Labels — legíveis, escuros ── */
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] p,
section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] label,
section[data-testid="stSidebar"] .stWidgetLabel p {
    color: #1C2B4A !important;
    font-size: 11.5px !important;
    font-weight: 500 !important;
}

/* Radio */
section[data-testid="stSidebar"] div[role="radiogroup"] label p,
section[data-testid="stSidebar"] div[role="radiogroup"] label span,
section[data-testid="stSidebar"] div[role="radiogroup"] label {
    color: #1C2B4A !important;
    font-size: 12px !important;
    font-weight: 400 !important;
}

/* Inputs */
section[data-testid="stSidebar"] input {
    background: #FFFFFF !important;
    color: #1C2B4A !important;
    border: 1px solid #C8D0E4 !important;
    border-radius: 6px !important;
    font-size: 12px !important;
    padding: 6px 10px !important;
    transition: border-color 0.15s !important;
}
section[data-testid="stSidebar"] input:focus {
    border-color: #0054A6 !important;
    box-shadow: 0 0 0 3px rgba(0,84,166,0.12) !important;
}
section[data-testid="stSidebar"] input::placeholder {
    color: #9CAABE !important;
}

/* Select */
section[data-testid="stSidebar"] div[data-baseweb="select"] > div {
    background: #FFFFFF !important;
    border: 1px solid #C8D0E4 !important;
    border-radius: 6px !important;
    color: #1C2B4A !important;
}
section[data-testid="stSidebar"] div[data-baseweb="select"] > div:focus-within {
    border-color: #0054A6 !important;
    box-shadow: 0 0 0 3px rgba(0,84,166,0.12) !important;
}
section[data-testid="stSidebar"] div[data-baseweb="select"] input {
    color: #1C2B4A !important;
    background: transparent !important;
    border: none !important;
}

/* Tags multiselect */
section[data-testid="stSidebar"] span[data-baseweb="tag"] {
    background-color: #EBF2FF !important;
    border: 1px solid #BDD3F5 !important;
    border-radius: 4px !important;
}
section[data-testid="stSidebar"] span[data-baseweb="tag"] span {
    color: #003D80 !important;
    font-size: 11px !important;
    font-weight: 500 !important;
}
section[data-testid="stSidebar"] [data-baseweb="tag"] svg {
    fill: #5A8AC8 !important;
}

/* Divisor */
section[data-testid="stSidebar"] hr {
    border-color: #E8ECF2 !important;
    margin: 14px 0 !important;
}

/* File uploader */
section[data-testid="stSidebar"] [data-testid="stFileUploader"] {
    background: #F7F9FC !important;
    border: 1.5px dashed #C8D0E4 !important;
    border-radius: 8px !important;
}
section[data-testid="stSidebar"] [data-testid="stFileUploader"] p,
section[data-testid="stSidebar"] [data-testid="stFileUploader"] span,
section[data-testid="stSidebar"] [data-testid="stFileUploader"] small {
    color: #5A6A8A !important;
    font-size: 11.5px !important;
}

/* Alert sidebar */
section[data-testid="stSidebar"] .stAlert {
    background: #EBF2FF !important;
    border: 1px solid #BDD3F5 !important;
    border-radius: 6px !important;
    color: #003D80 !important;
    font-size: 11.5px !important;
}

/* ════════════════════════════════════════════
   ÁREA PRINCIPAL
════════════════════════════════════════════ */

h1, h2, h3 {
    color: #1C2B4A !important;
    font-weight: 700 !important;
}

/* ── Métricas — card branco faixa azul ── */
[data-testid="stMetric"] {
    background: #FFFFFF !important;
    border-radius: 8px !important;
    padding: 18px 20px !important;
    border: 1px solid #E8ECF2 !important;
    border-top: 3px solid #0054A6 !important;
    box-shadow: 0 1px 4px rgba(28,43,74,0.06) !important;
}
[data-testid="stMetric"] label,
[data-testid="stMetric"] [data-testid="stMetricLabel"] p {
    color: #5A6A8A !important;
    font-size: 10.5px !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.08em !important;
}
[data-testid="stMetric"] [data-testid="stMetricValue"] {
    color: #1C2B4A !important;
    font-size: 1.80rem !important;
    font-weight: 700 !important;
    line-height: 1.15 !important;
}

/* ── Botão primário ── */
.stButton > button {
    background-color: #0054A6 !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 6px !important;
    font-weight: 600 !important;
    font-size: 12.5px !important;
    padding: 9px 22px !important;
    letter-spacing: 0.01em !important;
    box-shadow: 0 1px 4px rgba(0,84,166,0.25) !important;
    transition: background-color 0.15s, box-shadow 0.15s !important;
}
.stButton > button:hover {
    background-color: #003D80 !important;
    box-shadow: 0 3px 10px rgba(0,84,166,0.35) !important;
}

/* ── Botão download — outline ── */
.stDownloadButton > button {
    background-color: #FFFFFF !important;
    color: #0054A6 !important;
    border: 1.5px solid #0054A6 !important;
    border-radius: 6px !important;
    font-weight: 600 !important;
    font-size: 12.5px !important;
    padding: 9px 22px !important;
    transition: all 0.15s !important;
}
.stDownloadButton > button:hover {
    background-color: #EBF2FF !important;
}

/* ── Abas — underline azul ── */
.stTabs [data-baseweb="tab-list"] {
    background: #FFFFFF !important;
    border-bottom: 2px solid #E8ECF2 !important;
    padding: 0 !important;
    gap: 0 !important;
    box-shadow: 0 1px 0 #E8ECF2 !important;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    border-radius: 0 !important;
    font-size: 12.5px !important;
    font-weight: 500 !important;
    color: #5A6A8A !important;
    padding: 11px 20px !important;
    border-bottom: 2px solid transparent !important;
    margin-bottom: -2px !important;
    transition: color 0.15s !important;
}
.stTabs [aria-selected="true"] {
    color: #0054A6 !important;
    font-weight: 600 !important;
    border-bottom: 2px solid #0054A6 !important;
    background: transparent !important;
}
.stTabs [data-baseweb="tab"]:hover {
    color: #0054A6 !important;
    background: #F4F8FF !important;
}

/* ── Tabelas ── */
[data-testid="stDataFrame"] {
    border-radius: 8px !important;
    border: 1px solid #E8ECF2 !important;
    overflow: hidden !important;
    box-shadow: 0 1px 4px rgba(28,43,74,0.05) !important;
}

/* ── Expander ── */
[data-testid="stExpander"] {
    background: #FFFFFF !important;
    border: 1px solid #E8ECF2 !important;
    border-radius: 8px !important;
    box-shadow: 0 1px 4px rgba(28,43,74,0.04) !important;
}

/* ── Divisores ── */
hr {
    border-color: #E8ECF2 !important;
    margin: 16px 0 !important;
}

/* ── Alertas ── */
.stAlert {
    border-radius: 6px !important;
    font-size: 12.5px !important;
}

/* ── Captions ── */
.stCaption, [data-testid="stCaptionContainer"] p {
    color: #5A6A8A !important;
    font-size: 11px !important;
}

/* ── Card de gráfico ── */
.g-card {
    background: #FFFFFF;
    border: 1px solid #E8ECF2;
    border-radius: 8px;
    padding: 18px 20px 12px 20px;
    margin-bottom: 14px;
    box-shadow: 0 1px 4px rgba(28,43,74,0.05);
}
.g-label {
    font-size: 10.5px;
    font-weight: 700;
    color: #5A6A8A;
    text-transform: uppercase;
    letter-spacing: 0.09em;
    margin-bottom: 12px;
    padding-bottom: 10px;
    border-bottom: 1px solid #F0F3F8;
}

/* ── Sidebar — seção e marca ── */
.sb-brand {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 16px 0 14px 0;
    border-bottom: 1px solid #E8ECF2;
    margin-bottom: 16px;
}
.sb-brand-icon {
    width: 32px;
    height: 32px;
    background: #0054A6;
    border-radius: 6px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 15px;
    color: #FFFFFF;
    font-weight: 700;
    flex-shrink: 0;
}
.sb-brand-name {
    font-size: 12.5px;
    font-weight: 600;
    color: #1C2B4A;
    line-height: 1.2;
}
.sb-brand-sub {
    font-size: 10.5px;
    color: #5A6A8A;
    margin-top: 1px;
}
.sb-section {
    font-size: 10px;
    font-weight: 600;
    color: #9CAABE;
    text-transform: uppercase;
    letter-spacing: 0.11em;
    margin: 18px 0 8px 0;
}

/* ── Painel KPI ── */
.kpi-panel {
    background: #FFFFFF;
    border: 1px solid #E8ECF2;
    border-radius: 8px;
    padding: 18px 20px 14px 20px;
    margin-bottom: 18px;
    box-shadow: 0 1px 4px rgba(28,43,74,0.05);
}
.kpi-panel-title {
    font-size: 10.5px;
    font-weight: 700;
    color: #5A6A8A;
    text-transform: uppercase;
    letter-spacing: 0.09em;
    margin-bottom: 14px;
    padding-bottom: 10px;
    border-bottom: 1px solid #F0F3F8;
}

/* ── Breadcrumb ── */
.breadcrumb {
    font-size: 11px;
    color: #9CAABE;
    margin: 14px 0 6px 0;
    display: flex;
    align-items: center;
    gap: 6px;
}
.breadcrumb-sep { color: #C8D0E4; }
.breadcrumb-current { color: #1C2B4A; font-weight: 500; }

</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ──────────────────────────────────────────────────────────────────────────────
COLUNAS_OBRIGATORIAS = [
    "PEDIDO", "CLIENTE", "ESTADO", "REGIÃO",
    "FATURAR EM", "ITEM", "QNTD PROGRAMADA", "ESTOQUE INICIAL"
]

C_BLUE  = "#0054A6"
C_BLUE2 = "#1A73C8"
C_LGRAY = "#E8ECF2"
C_RED   = "#C0392B"
C_AMBER = "#B85C00"
C_GREEN = "#217A3C"
C_TEXT  = "#1C2B4A"
C_MUTED = "#5A6A8A"

# ──────────────────────────────────────────────────────────────────────────────
# FUNÇÕES DE NEGÓCIO
# ──────────────────────────────────────────────────────────────────────────────

def preparar_base(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.str.strip().str.upper()
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].astype(str).str.strip()
    df["FATURAR EM"]      = pd.to_datetime(df["FATURAR EM"], errors="coerce")
    df["QNTD PROGRAMADA"] = pd.to_numeric(df["QNTD PROGRAMADA"], errors="coerce").fillna(0)
    df["ESTOQUE INICIAL"] = pd.to_numeric(df["ESTOQUE INICIAL"], errors="coerce").fillna(0)
    return df

def validar_colunas(df: pd.DataFrame) -> list:
    return [c for c in COLUNAS_OBRIGATORIAS if c not in df.columns]

def montar_estoque(df: pd.DataFrame) -> dict:
    estoque = {}
    for _, row in df.iterrows():
        if row["ITEM"] not in estoque:
            estoque[row["ITEM"]] = row["ESTOQUE INICIAL"]
    return estoque

def aplicar_filtros(df, regioes, estados, clientes):
    if regioes:  df = df[df["REGIÃO"].isin(regioes)]
    if estados:  df = df[df["ESTADO"].isin(estados)]
    if clientes: df = df[df["CLIENTE"].isin(clientes)]
    return df

def ordenar_prioridade(df, p_regiao, p_estado):
    df = df.copy()
    p1 = set(df[df["REGIÃO"].isin(p_regiao)]["PEDIDO"].unique()) if p_regiao else set()
    p2 = set(df[df["ESTADO"].isin(p_estado) & ~df["PEDIDO"].isin(p1)]["PEDIDO"].unique()) if p_estado else set()
    df["ORDEM_PRIORIDADE"] = 3
    df.loc[df["PEDIDO"].isin(p1), "ORDEM_PRIORIDADE"] = 1
    df.loc[df["PEDIDO"].isin(p2), "ORDEM_PRIORIDADE"] = 2
    return df.sort_values(["ORDEM_PRIORIDADE", "FATURAR EM", "PEDIDO"], ascending=True)

def analisar_pedidos(df_ordem, estoque):
    liberados, bloqueados = [], []
    consumo_item, faltas_item = {}, {}

    for pedido, grupo in df_ordem.groupby("PEDIDO", sort=False):
        info   = grupo.iloc[0]
        faltas = []
        pode   = True

        for _, row in grupo.iterrows():
            item  = row["ITEM"]
            qtd   = row["QNTD PROGRAMADA"]
            saldo = estoque.get(item, 0)
            if saldo < qtd:
                pode = False
                faltas.append({"ITEM": item, "QTD": qtd, "SALDO": saldo, "FALTA": qtd - saldo})

        base = {
            "PEDIDO":     pedido,
            "CLIENTE":    info["CLIENTE"],
            "ESTADO":     info["ESTADO"],
            "REGIÃO":     info["REGIÃO"],
            "FATURAR EM": info["FATURAR EM"].date() if pd.notnull(info["FATURAR EM"]) else None,
            "PRIORIDADE": info["ORDEM_PRIORIDADE"],
        }

        if pode:
            itens = []
            for _, row in grupo.iterrows():
                item, qtd = row["ITEM"], row["QNTD PROGRAMADA"]
                estoque[item] = estoque.get(item, 0) - qtd
                consumo_item[item] = consumo_item.get(item, 0) + qtd
                itens.append(f"{item} ({int(qtd)})")
            liberados.append({**base, "ITENS": " | ".join(itens)})
        else:
            for i, f in enumerate(faltas, 1):
                base[f"ITEM_{i}"]               = f["ITEM"]
                base[f"QTD_NECESSARIA_{i}"]     = f["QTD"]
                base[f"ESTOQUE_DISPONIVEL_{i}"] = f["SALDO"]
                base[f"FALTA_{i}"]              = f["FALTA"]
                faltas_item[f["ITEM"]] = faltas_item.get(f["ITEM"], 0) + f["FALTA"]
            bloqueados.append(base)

    _ec = pd.DataFrame(columns=["ITEM", "CONSUMO_TOTAL"])
    _ef = pd.DataFrame(columns=["ITEM", "FALTA_TOTAL"])
    df_cons = pd.DataFrame([{"ITEM": k, "CONSUMO_TOTAL": v} for k, v in consumo_item.items()]
                           ).sort_values("CONSUMO_TOTAL", ascending=False) if consumo_item else _ec
    df_falt = pd.DataFrame([{"ITEM": k, "FALTA_TOTAL": v}   for k, v in faltas_item.items()]
                           ).sort_values("FALTA_TOTAL",   ascending=False) if faltas_item else _ef
    df_est  = pd.DataFrame([{"ITEM": k, "ESTOQUE_FINAL": v}  for k, v in estoque.items()]
                           ).sort_values("ESTOQUE_FINAL",  ascending=False)
    return pd.DataFrame(liberados), pd.DataFrame(bloqueados), df_cons, df_falt, df_est

def calcular_resumo(df_lib, df_bloq, df_cons, df_falt):
    total = len(df_lib) + len(df_bloq)
    pct   = round(len(df_lib) / total * 100, 1) if total else 0
    return pd.DataFrame([
        {"METRICA": "Total pedidos analisados", "VALOR": total},
        {"METRICA": "Pedidos liberados",         "VALOR": len(df_lib)},
        {"METRICA": "Pedidos bloqueados",         "VALOR": len(df_bloq)},
        {"METRICA": "Taxa de liberacao (%)",      "VALOR": pct},
        {"METRICA": "Total itens consumidos",     "VALOR": int(df_cons["CONSUMO_TOTAL"].sum()) if not df_cons.empty else 0},
        {"METRICA": "Total itens em falta",       "VALOR": int(df_falt["FALTA_TOTAL"].sum())   if not df_falt.empty else 0},
    ])

# ──────────────────────────────────────────────────────────────────────────────
# EXCEL
# ──────────────────────────────────────────────────────────────────────────────

def gerar_excel(df_lib, df_bloq, df_cons, df_falt, df_est, df_resumo) -> bytes:
    output = BytesIO()
    SHEETS = {
        "RESUMO":        (df_resumo, "0054A6"),
        "LIBERADOS":     (df_lib,    "217A3C"),
        "NAO_LIBERADOS": (df_bloq,   "C0392B"),
        "CONSUMO_ITEM":  (df_cons,   "0054A6"),
        "FALTAS_ITEM":   (df_falt,   "C0392B"),
        "ESTOQUE_FINAL": (df_est,    "217A3C"),
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for nome, (df, cor) in SHEETS.items():
            df.to_excel(writer, sheet_name=nome, index=False)
            ws = writer.sheets[nome]
            for cell in ws[1]:
                cell.fill      = PatternFill("solid", fgColor=cor)
                cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = Border(bottom=Side(style="thin", color="FFFFFF"))
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
                fc = "F4F6FA" if i % 2 == 0 else "FFFFFF"
                for cell in row:
                    cell.fill      = PatternFill("solid", fgColor=fc)
                    cell.font      = Font(name="Calibri", size=10)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            for col in ws.columns:
                letra = get_column_letter(col[0].column)
                ml = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[letra].width = min(max(ml + 3, 12), 50)
            ws.freeze_panes = "A2"
            if ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions
    return output.getvalue()

# ──────────────────────────────────────────────────────────────────────────────
# GRÁFICOS
# ──────────────────────────────────────────────────────────────────────────────

_L = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="Inter, system-ui, sans-serif", size=11, color=C_TEXT),

)

def fig_donut(liberados, bloqueados):
    total = liberados + bloqueados
    pct   = round(liberados / total * 100, 1) if total else 0
    fig   = go.Figure(go.Pie(
        labels=["Liberados", "Bloqueados"],
        values=[liberados, bloqueados],
        hole=0.72,
        marker=dict(colors=[C_BLUE, "#E4EAF5"], line=dict(color="#FFFFFF", width=2)),
        textinfo="none",
        hovertemplate="%{label}: %{value} (%{percent})<extra></extra>",
    ))
    fig.add_annotation(
        text=f"<b style='font-size:24px;color:{C_BLUE}'>{pct}%</b><br>"
             f"<span style='font-size:11px;color:{C_MUTED}'>liberado</span>",
        x=0.5, y=0.5, showarrow=False, align="center"
    )
    fig.update_layout(
        **_L, margin=dict(t=6, b=6, l=6, r=6), height=215,
        showlegend=True,
        legend=dict(orientation="h", x=0.5, xanchor="center", y=-0.06,
                    font=dict(size=11, color=C_MUTED)),
    )
    return fig

def fig_regiao(df_lib, df_bloq):
    regioes  = sorted(set(list(df_lib["REGIÃO"].unique()) + list(df_bloq["REGIÃO"].unique())))
    lib_cnt  = [len(df_lib[df_lib["REGIÃO"]   == r]) for r in regioes]
    bloq_cnt = [len(df_bloq[df_bloq["REGIÃO"] == r]) for r in regioes]
    fig = go.Figure()
    fig.add_trace(go.Bar(
        name="Liberados",  x=regioes, y=lib_cnt, marker_color=C_BLUE,
        text=lib_cnt, textposition="outside", textfont=dict(size=11, color=C_TEXT),
    ))
    fig.add_trace(go.Bar(
        name="Bloqueados", x=regioes, y=bloq_cnt, marker_color="#C5D3E8",
        text=bloq_cnt, textposition="outside", textfont=dict(size=11, color=C_MUTED),
    ))
    fig.update_layout(
        **_L, barmode="group", height=260,
        margin=dict(t=6, b=6, l=6, r=8),
        xaxis=dict(showgrid=False, tickfont=dict(size=11, color=C_TEXT),
                   linecolor=C_LGRAY, linewidth=1),
        yaxis=dict(showgrid=True, gridcolor="#F0F3F8", zeroline=False,
                   tickfont=dict(size=10, color=C_MUTED)),
        legend=dict(orientation="h", x=1, xanchor="right", y=1.12,
                    font=dict(size=11, color=C_MUTED), bgcolor="rgba(0,0,0,0)"),
    )
    return fig

def fig_barras_h(df, col_label, col_valor, cor, n=10):
    df_p = df.nlargest(n, col_valor).sort_values(col_valor, ascending=True)
    fig  = go.Figure(go.Bar(
        y=df_p[col_label].astype(str), x=df_p[col_valor],
        orientation="h", marker_color=cor, marker_line=dict(width=0),
        text=df_p[col_valor].apply(lambda v: f"{int(v):,}"),
        textposition="outside", textfont=dict(size=11, color=C_TEXT),
        hovertemplate="%{y}: %{x:,}<extra></extra>",
    ))
    fig.update_layout(
        **_L, height=max(190, len(df_p) * 32),
        margin=dict(t=6, b=6, l=6, r=56),
        xaxis=dict(showgrid=True, gridcolor="#F0F3F8", zeroline=False,
                   tickfont=dict(size=10, color=C_MUTED)),
        yaxis=dict(showgrid=False, tickfont=dict(size=11, color=C_TEXT)),
    )
    return fig

def fig_estoque(df_est, n=12):
    df_p  = df_est.nlargest(n, "ESTOQUE_FINAL").sort_values("ESTOQUE_FINAL", ascending=True)
    cores = [C_RED if v == 0 else C_AMBER if v < 20 else C_BLUE for v in df_p["ESTOQUE_FINAL"]]
    fig   = go.Figure(go.Bar(
        y=df_p["ITEM"].astype(str), x=df_p["ESTOQUE_FINAL"],
        orientation="h", marker_color=cores, marker_line=dict(width=0),
        text=df_p["ESTOQUE_FINAL"].apply(lambda v: f"{int(v):,}"),
        textposition="outside", textfont=dict(size=11, color=C_TEXT),
        hovertemplate="%{y}: %{x:,}<extra></extra>",
    ))
    fig.update_layout(
        **_L, height=max(190, len(df_p) * 32),
        margin=dict(t=6, b=6, l=6, r=56),
        xaxis=dict(showgrid=True, gridcolor="#F0F3F8", zeroline=False,
                   tickfont=dict(size=10, color=C_MUTED)),
        yaxis=dict(showgrid=False, tickfont=dict(size=11, color=C_TEXT)),
    )
    return fig

# ──────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ──────────────────────────────────────────────────────────────────────────────

with st.sidebar:

    st.markdown("""
    <div class="sb-brand">
        <div class="sb-brand-icon">LP</div>
        <div>
            <div class="sb-brand-name">Liberacao de Pedidos</div>
            <div class="sb-brand-sub">Simulacao por Estoque</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    arquivo = st.file_uploader("Planilha base (.xlsx)", type=["xlsx"])
    if arquivo:
        st.success("Planilha carregada.")

    st.markdown('<div class="sb-section">Periodo de Analise</div>', unsafe_allow_html=True)
    tipo_data = st.radio("Tipo", ["Data especifica", "Periodo"], label_visibility="collapsed")
    if tipo_data == "Data especifica":
        data_ref    = st.date_input("Data", value=date.today(), label_visibility="collapsed")
        data_inicio = data_fim = data_ref
    else:
        data_inicio = st.date_input("Inicio", value=date.today())
        data_fim    = st.date_input("Fim",    value=date.today())

    opcoes_regiao = opcoes_estado = opcoes_cliente = []
    df_global = None
    if arquivo:
        try:
            df_tmp = preparar_base(pd.read_excel(arquivo))
            if not validar_colunas(df_tmp):
                df_global      = df_tmp
                opcoes_regiao  = sorted(df_tmp["REGIÃO"].dropna().unique())
                opcoes_estado  = sorted(df_tmp["ESTADO"].dropna().unique())
                opcoes_cliente = sorted(df_tmp["CLIENTE"].dropna().unique())
        except Exception:
            pass

    st.markdown('<div class="sb-section">Prioridade de Liberacao</div>', unsafe_allow_html=True)
    st.caption("P1 = regioes  ·  P2 = estados  ·  P3 = demais")
    p_regiao = st.multiselect("Regioes (P1)", opcoes_regiao, placeholder="Nenhuma")
    p_estado = st.multiselect("Estados (P2)",  opcoes_estado,  placeholder="Nenhum")

    st.markdown('<div class="sb-section">Filtros Operacionais</div>', unsafe_allow_html=True)
    f_regiao  = st.multiselect("Regiao",  opcoes_regiao,  placeholder="Todas")
    f_estado  = st.multiselect("Estado",  opcoes_estado,  placeholder="Todos")
    f_cliente = st.multiselect("Cliente", opcoes_cliente, placeholder="Todos")

    st.markdown("---")
    rodar = st.button("Gerar Analise", use_container_width=True)

# ──────────────────────────────────────────────────────────────────────────────
# TOPBAR
# ──────────────────────────────────────────────────────────────────────────────

st.markdown(f"""
<div class="totvs-topbar">
    <div class="totvs-topbar-left">
        <div class="totvs-logo">Gestao <span>Operacional</span></div>
        <div class="totvs-divider"></div>
        <div class="totvs-page-name">Planejamento de Liberacao de Pedidos</div>
    </div>
    <div class="totvs-topbar-right">{date.today().strftime('%d/%m/%Y')}</div>
</div>
""", unsafe_allow_html=True)

# Breadcrumb
st.markdown("""
<div class="breadcrumb">
    Estoque <span class="breadcrumb-sep">›</span>
    Simulacao <span class="breadcrumb-sep">›</span>
    <span class="breadcrumb-current">Liberacao de Pedidos</span>
</div>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# VALIDAÇÕES
# ──────────────────────────────────────────────────────────────────────────────

if not arquivo:
    st.info("Carregue a planilha base na barra lateral para iniciar a analise.")
    st.stop()

if df_global is None:
    erros = validar_colunas(preparar_base(pd.read_excel(arquivo)))
    st.error(f"Colunas obrigatorias ausentes: {', '.join(erros)}")
    st.caption("Esperado: " + "  ·  ".join(COLUNAS_OBRIGATORIAS))
    st.stop()

df = df_global

with st.expander("Visualizar base carregada", expanded=False):
    st.dataframe(df.head(100), use_container_width=True, hide_index=True)
    st.caption(
        f"{len(df):,} linhas  ·  {df['PEDIDO'].nunique():,} pedidos  "
        f"·  {df['ITEM'].nunique():,} itens  ·  {df['CLIENTE'].nunique():,} clientes"
    )

if not rodar:
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Linhas na base",  f"{len(df):,}")
    c2.metric("Pedidos unicos",  f"{df['PEDIDO'].nunique():,}")
    c3.metric("Itens unicos",    f"{df['ITEM'].nunique():,}")
    c4.metric("Clientes unicos", f"{df['CLIENTE'].nunique():,}")
    st.stop()

# ──────────────────────────────────────────────────────────────────────────────
# PROCESSAMENTO
# ──────────────────────────────────────────────────────────────────────────────

with st.spinner("Processando..."):
    df_f = df[
        (df["FATURAR EM"].dt.date >= data_inicio) &
        (df["FATURAR EM"].dt.date <= data_fim)
    ]
    if df_f.empty:
        st.warning("Nenhum pedido no periodo selecionado.")
        st.stop()

    df_f = aplicar_filtros(df_f, f_regiao, f_estado, f_cliente)
    if df_f.empty:
        st.warning("Nenhum pedido com os filtros aplicados.")
        st.stop()

    estoque   = montar_estoque(df)
    df_ordem  = ordenar_prioridade(df_f, p_regiao, p_estado)
    df_lib, df_bloq, df_cons, df_falt, df_est = analisar_pedidos(df_ordem, estoque)
    df_resumo = calcular_resumo(df_lib, df_bloq, df_cons, df_falt)

# ──────────────────────────────────────────────────────────────────────────────
# KPIs
# ──────────────────────────────────────────────────────────────────────────────

total = len(df_lib) + len(df_bloq)
pct   = round(len(df_lib) / total * 100, 1) if total else 0

st.markdown('<div class="kpi-panel"><div class="kpi-panel-title">Resultado Geral da Simulacao</div>', unsafe_allow_html=True)
c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("Pedidos Analisados",     f"{total:,}")
c2.metric("Pedidos Liberados",      f"{len(df_lib):,}")
c3.metric("Pedidos Bloqueados",     f"{len(df_bloq):,}")
c4.metric("Taxa Liberacao", f"{pct:.1f}%")
c5.metric("Total de SKUs em Falta", f"{df_falt['ITEM'].nunique() if not df_falt.empty else 0:,}")
st.markdown('</div>', unsafe_allow_html=True)

st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# ABAS
# ──────────────────────────────────────────────────────────────────────────────

tab_dash, tab_lib, tab_bloq, tab_falt, tab_cons, tab_est = st.tabs([
    "Visao Gerencial",
    "Liberados",
    "Nao Liberados",
    "Faltas por Item",
    "Consumo por Item",
    "Estoque Final",
])

# ── VISÃO GERENCIAL ──────────────────────────────────────────────────────────
with tab_dash:
    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

    col_a, col_b = st.columns([1, 2])

    with col_a:
        st.markdown('<div class="g-card"><div class="g-label">Taxa de Liberacao</div>', unsafe_allow_html=True)
        st.plotly_chart(fig_donut(len(df_lib), len(df_bloq)),
                        use_container_width=True, config={"displayModeBar": False})
        st.markdown('</div>', unsafe_allow_html=True)

    with col_b:
        st.markdown('<div class="g-card"><div class="g-label">Liberados vs Bloqueados por Regiao</div>', unsafe_allow_html=True)
        _lr = df_lib  if not df_lib.empty  else pd.DataFrame(columns=["REGIÃO"])
        _br = df_bloq if not df_bloq.empty else pd.DataFrame(columns=["REGIÃO"])
        st.plotly_chart(fig_regiao(_lr, _br),
                        use_container_width=True, config={"displayModeBar": False})
        st.markdown('</div>', unsafe_allow_html=True)

# ── LIBERADOS ────────────────────────────────────────────────────────────────
with tab_lib:
    if df_lib.empty:
        st.warning("Nenhum pedido foi liberado.")
    else:
        st.caption(f"{len(df_lib):,} pedidos liberados")
        st.dataframe(df_lib, use_container_width=True, hide_index=True)

# ── NAO LIBERADOS ────────────────────────────────────────────────────────────
with tab_bloq:
    if df_bloq.empty:
        st.success("Todos os pedidos foram liberados.")
    else:
        st.caption(f"{len(df_bloq):,} pedidos bloqueados por insuficiencia de estoque")
        st.dataframe(df_bloq, use_container_width=True, hide_index=True)

# ── FALTAS ───────────────────────────────────────────────────────────────────
with tab_falt:
    if df_falt.empty:
        st.success("Nenhum item em falta.")
    else:
        st.caption(f"{len(df_falt):,} itens com falta registrada")
        st.dataframe(df_falt, use_container_width=True, hide_index=True)

# ── CONSUMO ──────────────────────────────────────────────────────────────────
with tab_cons:
    if df_cons.empty:
        st.info("Sem consumo registrado.")
    else:
        st.caption(f"{len(df_cons):,} itens consumidos")
        st.dataframe(df_cons, use_container_width=True, hide_index=True)

# ── ESTOQUE FINAL ─────────────────────────────────────────────────────────────
with tab_est:
    st.caption("Saldo simulado por item apos liberacao dos pedidos possiveis")
    st.dataframe(df_est, use_container_width=True, hide_index=True)

# ──────────────────────────────────────────────────────────────────────────────
# EXPORTAR
# ──────────────────────────────────────────────────────────────────────────────

st.markdown("---")
col_dl, col_txt = st.columns([1, 3])
with col_dl:
    excel_bytes = gerar_excel(df_lib, df_bloq, df_cons, df_falt, df_est, df_resumo)
    st.download_button(
        label="Baixar Relatorio Excel",
        data=excel_bytes,
        file_name=f"liberacao_{date.today().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with col_txt:
    st.caption(
        f"6 abas: Resumo · Liberados · Nao Liberados · Consumo · Faltas · Estoque Final"
        f"  ·  Gerado em {date.today().strftime('%d/%m/%Y')}"
    )
