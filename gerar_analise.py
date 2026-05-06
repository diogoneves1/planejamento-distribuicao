import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import date

print("Iniciando análise...")

# ──────────────────────────────────────────────────────────────────────────────
# 1. LER E PADRONIZAR
# ──────────────────────────────────────────────────────────────────────────────
df = pd.read_excel("base.xlsx")
df.columns = df.columns.str.strip().str.upper()

print("Colunas encontradas:", df.columns.tolist())

df["FATURAR EM"]      = pd.to_datetime(df["FATURAR EM"], errors="coerce").dt.date
df["QNTD PROGRAMADA"] = pd.to_numeric(df["QNTD PROGRAMADA"], errors="coerce").fillna(0)
df["ESTOQUE INICIAL"] = pd.to_numeric(df["ESTOQUE INICIAL"], errors="coerce").fillna(0)

hoje = date.today()

# ──────────────────────────────────────────────────────────────────────────────
# 2. FILTRAR (hoje + atrasados)
# ──────────────────────────────────────────────────────────────────────────────
df_filtrado = df[df["FATURAR EM"] <= hoje].copy()

# ──────────────────────────────────────────────────────────────────────────────
# 3. MONTAR ESTOQUE VIRTUAL
# ──────────────────────────────────────────────────────────────────────────────
estoque = {}
for _, row in df.iterrows():
    item = row["ITEM"]
    if item not in estoque:
        estoque[item] = row["ESTOQUE INICIAL"]

# ──────────────────────────────────────────────────────────────────────────────
# 4. PROCESSAR PEDIDOS
# ──────────────────────────────────────────────────────────────────────────────
liberados     = []
nao_liberados = []
consumo_item  = {}
faltas_item   = {}

for pedido, grupo in df_filtrado.groupby("PEDIDO"):
    info  = grupo.iloc[0]
    faltas = []
    pode  = True

    for _, row in grupo.iterrows():
        item  = row["ITEM"]
        qtd   = row["QNTD PROGRAMADA"]
        saldo = estoque.get(item, 0)
        if saldo < qtd:
            pode = False
            faltas.append({
                "ITEM": item, "QTD": qtd,
                "SALDO": saldo, "FALTA": qtd - saldo
            })

    base = {
        "PEDIDO":     pedido,
        "CLIENTE":    info.get("CLIENTE", ""),
        "ESTADO":     info.get("ESTADO", ""),
        "REGIAO":     info.get("REGIÃO", ""),
        "FATURAR EM": info["FATURAR EM"],
    }

    if pode:
        itens = []
        for _, row in grupo.iterrows():
            item, qtd = row["ITEM"], row["QNTD PROGRAMADA"]
            estoque[item] = estoque.get(item, 0) - qtd
            consumo_item[item] = consumo_item.get(item, 0) + qtd
            itens.append(f"{item} ({int(qtd)})")
        liberados.append({**base, "ITENS": " | ".join(itens)})
    else:
        for i, f in enumerate(faltas, 1):
            base[f"ITEM_{i}"]    = f["ITEM"]
            base[f"QTD_{i}"]     = f["QTD"]
            base[f"SALDO_{i}"]   = f["SALDO"]
            base[f"FALTA_{i}"]   = f["FALTA"]
            faltas_item[f["ITEM"]] = faltas_item.get(f["ITEM"], 0) + f["FALTA"]
        nao_liberados.append(base)

# ──────────────────────────────────────────────────────────────────────────────
# 5. MONTAR DataFrames DE RESULTADO
# ──────────────────────────────────────────────────────────────────────────────
df_lib  = pd.DataFrame(liberados)
df_bloq = pd.DataFrame(nao_liberados)

df_cons = pd.DataFrame(
    [{"ITEM": k, "CONSUMO_TOTAL": v} for k, v in consumo_item.items()]
).sort_values("CONSUMO_TOTAL", ascending=False) if consumo_item else pd.DataFrame(columns=["ITEM", "CONSUMO_TOTAL"])

df_falt = pd.DataFrame(
    [{"ITEM": k, "FALTA_TOTAL": v} for k, v in faltas_item.items()]
).sort_values("FALTA_TOTAL", ascending=False) if faltas_item else pd.DataFrame(columns=["ITEM", "FALTA_TOTAL"])

df_est = pd.DataFrame(
    [{"ITEM": k, "ESTOQUE_FINAL": v} for k, v in estoque.items()]
).sort_values("ESTOQUE_FINAL", ascending=False)

# ──────────────────────────────────────────────────────────────────────────────
# 6. RESUMO GERENCIAL
# ──────────────────────────────────────────────────────────────────────────────
total     = len(df_lib) + len(df_bloq)
pct_lib   = round(len(df_lib) / total * 100, 1) if total else 0
tot_cons  = int(df_cons["CONSUMO_TOTAL"].sum()) if not df_cons.empty else 0
tot_falt  = int(df_falt["FALTA_TOTAL"].sum())   if not df_falt.empty else 0

df_resumo = pd.DataFrame([
    {"METRICA": "Data de geracao",           "VALOR": str(hoje)},
    {"METRICA": "Total pedidos analisados",  "VALOR": total},
    {"METRICA": "Pedidos liberados",         "VALOR": len(df_lib)},
    {"METRICA": "Pedidos bloqueados",        "VALOR": len(df_bloq)},
    {"METRICA": "Taxa de liberacao (%)",     "VALOR": pct_lib},
    {"METRICA": "Total itens consumidos",    "VALOR": tot_cons},
    {"METRICA": "Total itens em falta",      "VALOR": tot_falt},
])

# resumo por região (para gráfico)
df_reg_lib  = df_lib["REGIAO"].value_counts().reset_index().rename(columns={"index": "REGIAO", "REGIAO": "LIBERADOS"})  if not df_lib.empty  else pd.DataFrame(columns=["REGIAO", "LIBERADOS"])
df_reg_bloq = df_bloq["REGIAO"].value_counts().reset_index().rename(columns={"index": "REGIAO", "REGIAO": "BLOQUEADOS"}) if not df_bloq.empty else pd.DataFrame(columns=["REGIAO", "BLOQUEADOS"])
df_reg_lib.columns  = ["REGIAO", "LIBERADOS"]
df_reg_bloq.columns = ["REGIAO", "BLOQUEADOS"]
df_regiao = pd.merge(df_reg_lib, df_reg_bloq, on="REGIAO", how="outer").fillna(0)
df_regiao[["LIBERADOS", "BLOQUEADOS"]] = df_regiao[["LIBERADOS", "BLOQUEADOS"]].astype(int)

print(f"\nResultado:")
print(f"  Pedidos analisados : {total}")
print(f"  Liberados          : {len(df_lib)}")
print(f"  Bloqueados         : {len(df_bloq)}")
print(f"  Taxa de liberacao  : {pct_lib}%")

# ──────────────────────────────────────────────────────────────────────────────
# 7. EXPORTAR EXCEL
# ──────────────────────────────────────────────────────────────────────────────
ARQUIVO = "resultado.xlsx"

COR_NAVY   = "0054A6"
COR_VERDE  = "217A3C"
COR_VERM   = "C0392B"
COR_CINZA  = "F4F6FA"
COR_BRANCO = "FFFFFF"

def estilo_cabecalho(ws, cor=COR_NAVY):
    for cell in ws[1]:
        cell.fill      = PatternFill("solid", fgColor=cor)
        cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(bottom=Side(style="thin", color="FFFFFF"))

def estilo_dados(ws):
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
        fc = COR_CINZA if i % 2 == 0 else COR_BRANCO
        for cell in row:
            cell.fill      = PatternFill("solid", fgColor=fc)
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")

def ajustar_largura(ws):
    for col in ws.columns:
        letra = get_column_letter(col[0].column)
        ml = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[letra].width = min(max(ml + 3, 12), 50)

SHEETS = {
    "LIBERADOS":     (df_lib,    COR_VERDE),
    "NAO_LIBERADOS": (df_bloq,   COR_VERM),
    "CONSUMO_ITEM":  (df_cons,   COR_NAVY),
    "FALTAS_ITEM":   (df_falt,   COR_VERM),
    "ESTOQUE_FINAL": (df_est,    COR_VERDE),
    "RESUMO":        (df_resumo, COR_NAVY),
    "POR_REGIAO":    (df_regiao, COR_NAVY),
}

with pd.ExcelWriter(ARQUIVO, engine="openpyxl") as writer:
    for nome, (df_s, cor) in SHEETS.items():
        df_s.to_excel(writer, sheet_name=nome, index=False)
        ws = writer.sheets[nome]
        estilo_cabecalho(ws, cor)
        estilo_dados(ws)
        ajustar_largura(ws)
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

# ──────────────────────────────────────────────────────────────────────────────
# 8. VISÃO GERENCIAL — aba com KPIs + gráficos
# ──────────────────────────────────────────────────────────────────────────────
wb = load_workbook(ARQUIVO)

# ── Criar aba VISAO_GERENCIAL ──
if "VISAO_GERENCIAL" in wb.sheetnames:
    del wb["VISAO_GERENCIAL"]

ws_vg = wb.create_sheet("VISAO_GERENCIAL", 0)   # primeira aba

# Paletas de preenchimento
fill_navy   = PatternFill("solid", fgColor=COR_NAVY)
fill_verde  = PatternFill("solid", fgColor=COR_VERDE)
fill_verm   = PatternFill("solid", fgColor=COR_VERM)
fill_cinza  = PatternFill("solid", fgColor="F4F6FA")
fill_branco = PatternFill("solid", fgColor="FFFFFF")
fill_titulo = PatternFill("solid", fgColor="002D6B")

fonte_branca_g  = Font(name="Calibri", bold=True, color="FFFFFF", size=18)
fonte_branca_m  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
fonte_branca_p  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
fonte_navy_g    = Font(name="Calibri", bold=True, color="0054A6", size=22)
fonte_navy_m    = Font(name="Calibri", bold=True, color="0054A6", size=13)
fonte_verde_g   = Font(name="Calibri", bold=True, color="217A3C", size=22)
fonte_verm_g    = Font(name="Calibri", bold=True, color="C0392B", size=22)
fonte_label     = Font(name="Calibri", bold=True, color="5A6A8A", size=9)
fonte_normal    = Font(name="Calibri", size=10, color="1C2B4A")
fonte_titulo    = Font(name="Calibri", bold=True, color="FFFFFF", size=14)

centro = Alignment(horizontal="center", vertical="center")
esq    = Alignment(horizontal="left",   vertical="center")

borda_card = Border(
    top=Side(style="thin", color="E8ECF2"),
    left=Side(style="thin", color="E8ECF2"),
    right=Side(style="thin", color="E8ECF2"),
    bottom=Side(style="thin", color="E8ECF2"),
)

def preencher_range(ws, linha_ini, col_ini, linha_fim, col_fim, fill):
    for r in range(linha_ini, linha_fim + 1):
        for c in range(col_ini, col_fim + 1):
            ws.cell(r, c).fill = fill

def mesclar_e_escrever(ws, linha, col_ini, col_fim, valor, font, fill=None, align=None):
    ws.merge_cells(
        start_row=linha, start_column=col_ini,
        end_row=linha,   end_column=col_fim
    )
    cell = ws.cell(linha, col_ini)
    cell.value     = valor
    cell.font      = font
    cell.alignment = align or centro
    if fill:
        for c in range(col_ini, col_fim + 1):
            ws.cell(linha, c).fill = fill

# ── Largura das colunas ──
larguras = [1, 18, 3, 18, 3, 18, 3, 18, 3, 18, 1]
for i, w in enumerate(larguras, 1):
    ws_vg.column_dimensions[get_column_letter(i)].width = w

# Altura das linhas
alturas = {
    1: 8, 2: 30, 3: 8, 4: 14, 5: 28, 6: 8,
    7: 14, 8: 28, 9: 8, 10: 14, 11: 28, 12: 8,
    13: 14, 14: 28, 15: 8,
}
for ln, h in alturas.items():
    ws_vg.row_dimensions[ln].height = h

# ── Fundo geral ──
preencher_range(ws_vg, 1, 1, 50, 11, fill_cinza)

# ── Cabeçalho / título ──
preencher_range(ws_vg, 2, 1, 2, 11, fill_titulo)
mesclar_e_escrever(ws_vg, 2, 2, 8, "  VISAO GERENCIAL — LIBERACAO DE PEDIDOS", fonte_titulo, fill_titulo, esq)
ws_vg.cell(2, 10).value     = str(hoje.strftime("%d/%m/%Y"))
ws_vg.cell(2, 10).font      = Font(name="Calibri", color="FFFFFF", size=10)
ws_vg.cell(2, 10).alignment = Alignment(horizontal="right", vertical="center")
ws_vg.cell(2, 10).fill      = fill_titulo
ws_vg.cell(2, 11).fill      = fill_titulo

# ── Separador ──
preencher_range(ws_vg, 3, 1, 3, 11, PatternFill("solid", fgColor="0054A6"))

# ── Linha de rótulo dos KPIs ──
kpi_cols = [
    (2,  "PEDIDOS ANALISADOS"),
    (4,  "LIBERADOS"),
    (6,  "BLOQUEADOS"),
    (8,  "TAXA DE LIBERACAO"),
    (10, "ITENS EM FALTA"),
]
preencher_range(ws_vg, 4, 1, 4, 11, fill_branco)
for col, label in kpi_cols:
    ws_vg.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
    c = ws_vg.cell(4, col)
    c.value = label
    c.font  = fonte_label
    c.alignment = centro
    c.fill  = fill_branco

# ── Valores dos KPIs ──
kpi_vals = [
    (2,  total,           fonte_navy_g),
    (4,  len(df_lib),     fonte_verde_g),
    (6,  len(df_bloq),    fonte_verm_g),
    (8,  f"{pct_lib}%",   fonte_navy_g),
    (10, tot_falt,        fonte_verm_g),
]
preencher_range(ws_vg, 5, 1, 5, 11, fill_branco)
for col, val, font in kpi_vals:
    ws_vg.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
    c = ws_vg.cell(5, col)
    c.value     = val
    c.font      = font
    c.alignment = centro
    c.fill      = fill_branco

# Borda inferior do bloco KPI
preencher_range(ws_vg, 6, 1, 6, 11, PatternFill("solid", fgColor="E8ECF2"))

# ── Bloco: Top Faltas ──
preencher_range(ws_vg, 7, 2, 7, 5, fill_navy)
ws_vg.merge_cells(start_row=7, start_column=2, end_row=7, end_column=5)
c = ws_vg.cell(7, 2)
c.value = "TOP ITENS EM FALTA"
c.font  = fonte_branca_p
c.alignment = centro

# dados faltas (até 5 itens)
top_falt = df_falt.head(5).reset_index(drop=True)
for i in range(5):
    row_idx = 8 + i
    ws_vg.row_dimensions[row_idx].height = 16
    preencher_range(ws_vg, row_idx, 2, row_idx, 5, fill_branco)

    item_val  = top_falt.iloc[i]["ITEM"]       if i < len(top_falt) else ""
    falta_val = top_falt.iloc[i]["FALTA_TOTAL"] if i < len(top_falt) else ""

    c_item = ws_vg.cell(row_idx, 2)
    c_item.value = item_val
    c_item.font  = fonte_normal
    c_item.alignment = esq
    ws_vg.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)

    c_val = ws_vg.cell(row_idx, 4)
    c_val.value = int(falta_val) if falta_val != "" else ""
    c_val.font  = Font(name="Calibri", bold=True, size=10,
                       color="C0392B" if falta_val != "" else "FFFFFF")
    c_val.alignment = Alignment(horizontal="right", vertical="center")
    ws_vg.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx, end_column=5)

# ── Bloco: Top Consumo ──
preencher_range(ws_vg, 7, 7, 7, 10, fill_navy)
ws_vg.merge_cells(start_row=7, start_column=7, end_row=7, end_column=10)
c = ws_vg.cell(7, 7)
c.value = "TOP ITENS POR CONSUMO"
c.font  = fonte_branca_p
c.alignment = centro

top_cons = df_cons.head(5).reset_index(drop=True)
for i in range(5):
    row_idx = 8 + i
    preencher_range(ws_vg, row_idx, 7, row_idx, 10, fill_branco)

    item_val = top_cons.iloc[i]["ITEM"]         if i < len(top_cons) else ""
    cons_val = top_cons.iloc[i]["CONSUMO_TOTAL"] if i < len(top_cons) else ""

    c_item = ws_vg.cell(row_idx, 7)
    c_item.value = item_val
    c_item.font  = fonte_normal
    c_item.alignment = esq
    ws_vg.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=8)

    c_val = ws_vg.cell(row_idx, 9)
    c_val.value = int(cons_val) if cons_val != "" else ""
    c_val.font  = Font(name="Calibri", bold=True, size=10,
                       color="0054A6" if cons_val != "" else "FFFFFF")
    c_val.alignment = Alignment(horizontal="right", vertical="center")
    ws_vg.merge_cells(start_row=row_idx, start_column=9, end_row=row_idx, end_column=10)

# ── Separador ──
preencher_range(ws_vg, 14, 1, 14, 11, PatternFill("solid", fgColor="E8ECF2"))

# ── Bloco: Estoque Final ──
preencher_range(ws_vg, 15, 2, 15, 10, fill_navy)
ws_vg.merge_cells(start_row=15, start_column=2, end_row=15, end_column=10)
c = ws_vg.cell(15, 2)
c.value = "ESTOQUE FINAL SIMULADO POR ITEM (TOP 8)"
c.font  = fonte_branca_p
c.alignment = centro

top_est = df_est.head(8).reset_index(drop=True)
for i in range(8):
    row_idx = 16 + i
    ws_vg.row_dimensions[row_idx].height = 15
    preencher_range(ws_vg, row_idx, 2, row_idx, 10, fill_branco)

    item_val = top_est.iloc[i]["ITEM"]          if i < len(top_est) else ""
    est_val  = top_est.iloc[i]["ESTOQUE_FINAL"] if i < len(top_est) else ""

    c_item = ws_vg.cell(row_idx, 2)
    c_item.value = item_val
    c_item.font  = fonte_normal
    c_item.alignment = esq
    ws_vg.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)

    cor_val = "C0392B" if est_val == 0 else "B85C00" if est_val != "" and est_val < 20 else "0054A6"
    c_val = ws_vg.cell(row_idx, 7)
    c_val.value = int(est_val) if est_val != "" else ""
    c_val.font  = Font(name="Calibri", bold=True, size=10, color=cor_val)
    c_val.alignment = Alignment(horizontal="right", vertical="center")
    ws_vg.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=10)

# ── GRÁFICO 1: Pizza — Liberados vs Bloqueados ──
linha_grafico = 26

# dados auxiliares para o gráfico (escritos em área fora da view)
ws_vg.cell(48, 1).value = "Liberados"
ws_vg.cell(49, 1).value = "Bloqueados"
ws_vg.cell(48, 2).value = len(df_lib)
ws_vg.cell(49, 2).value = len(df_bloq)

pizza = PieChart()
pizza.title  = "Liberados vs Bloqueados"
pizza.style  = 10
pizza.width  = 12
pizza.height = 8

labels = Reference(ws_vg, min_col=1, min_row=48, max_row=49)
dados  = Reference(ws_vg, min_col=2, min_row=48, max_row=49)
pizza.add_data(dados)
pizza.set_categories(labels)
pizza.series[0].title = None

# colorir fatias
slice_lib  = DataPoint(idx=0)
slice_bloq = DataPoint(idx=1)
slice_lib.graphicalProperties.solidFill  = "0054A6"
slice_bloq.graphicalProperties.solidFill = "E8ECF2"
pizza.series[0].dPt = [slice_lib, slice_bloq]
pizza.series[0].dLbls = None

ws_vg.add_chart(pizza, f"B{linha_grafico}")

# ── GRÁFICO 2: Barras — Por Região ──
if not df_regiao.empty:
    linha_dados_reg = 48
    col_reg_ini = 4

    ws_vg.cell(linha_dados_reg,     col_reg_ini).value = "Regiao"
    ws_vg.cell(linha_dados_reg,     col_reg_ini + 1).value = "Liberados"
    ws_vg.cell(linha_dados_reg,     col_reg_ini + 2).value = "Bloqueados"

    for i, row in df_regiao.iterrows():
        ws_vg.cell(linha_dados_reg + 1 + i, col_reg_ini).value     = row["REGIAO"]
        ws_vg.cell(linha_dados_reg + 1 + i, col_reg_ini + 1).value = row["LIBERADOS"]
        ws_vg.cell(linha_dados_reg + 1 + i, col_reg_ini + 2).value = row["BLOQUEADOS"]

    n_reg = len(df_regiao)
    barras = BarChart()
    barras.type    = "col"
    barras.title   = "Pedidos por Regiao"
    barras.style   = 10
    barras.width   = 14
    barras.height  = 8
    barras.grouping = "clustered"

    cats = Reference(ws_vg, min_col=col_reg_ini,
                     min_row=linha_dados_reg + 1,
                     max_row=linha_dados_reg + n_reg)

    for col_offset, nome_serie in [(col_reg_ini + 1, "Liberados"), (col_reg_ini + 2, "Bloqueados")]:
        dados_reg = Reference(ws_vg, min_col=col_offset,
                              min_row=linha_dados_reg,
                              max_row=linha_dados_reg + n_reg)
        barras.add_data(dados_reg, titles_from_data=True)

    barras.set_categories(cats)
    barras.series[0].graphicalProperties.solidFill = "0054A6"
    barras.series[1].graphicalProperties.solidFill = "C5D3E8"

    ws_vg.add_chart(barras, f"F{linha_grafico}")

# ── Ocultar linhas auxiliares de dados dos gráficos ──
for r in range(48, 58):
    ws_vg.row_dimensions[r].hidden = True

# ── Título da seção de gráficos ──
ws_vg.row_dimensions[25].height = 14
preencher_range(ws_vg, 25, 2, 25, 10, fill_navy)
ws_vg.merge_cells(start_row=25, start_column=2, end_row=25, end_column=10)
c = ws_vg.cell(25, 2)
c.value = "GRAFICOS GERENCIAIS"
c.font  = fonte_branca_p
c.alignment = centro

# ── Proteger visuais — congelar painel ──
ws_vg.freeze_panes = "A3"

# ── Reordenar abas ──
nomes_ordem = [
    "VISAO_GERENCIAL", "LIBERADOS", "NAO_LIBERADOS",
    "CONSUMO_ITEM", "FALTAS_ITEM", "ESTOQUE_FINAL",
    "RESUMO", "POR_REGIAO"
]
for nome in nomes_ordem:
    if nome in wb.sheetnames:
        ws_tmp = wb[nome]
        wb.move_sheet(ws_tmp, offset=-len(wb.sheetnames))

wb.save(ARQUIVO)

print(f"\nArquivo gerado: {ARQUIVO}")
print("Abas: VISAO_GERENCIAL | LIBERADOS | NAO_LIBERADOS | CONSUMO_ITEM | FALTAS_ITEM | ESTOQUE_FINAL | RESUMO | POR_REGIAO")
print("Analise finalizada com sucesso.")
